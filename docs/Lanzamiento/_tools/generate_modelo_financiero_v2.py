#!/usr/bin/env python3
"""
Genera MODELO_FINANCIERO_ZONIX_PHARMA.xlsx (v3.8) — bottom-up + layout Detallado tipo Pizza QLQ.
Fuente: PROYECCION §1.1, PRESUPUESTO §2, ESTRUCTURA_LEGAL §1.4, benchmarks AVGH/Talently VE 2026.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "MODELO_FINANCIERO_ZONIX_PHARMA.xlsx"
DESCARGAS = Path("/home/aipp/Descargas/MODELO_FINANCIERO_ZONIX_PHARMA.xlsx")

MONTHS = list(range(1, 13))
DISCOUNT = 0.25

# Legal v2.4 (constitución C.A.)
LEGAL_SUBTOTAL = 5_050
INTRO_LEAN_BASE = 3_008
VIDEO_B2B_INTRO = 800  # one-shot video pitch farmacia (v3.1)
VIDEO_B2B_CASE = 800  # one-shot video caso éxito farmacia (v3.2 Base+)
INTRO_LEAN = INTRO_LEAN_BASE + VIDEO_B2B_INTRO
HQ_CAPEX_LEAN = 5_350
HOSTING_SAAS_M = 120
HOSTING_SMS_M = 34  # conservador; resto en contingencia

# Curva comercial original (4× Sales) — se escala por tier
_FIRMADAS_4S = [13, 13, 14, 15, 15, 16, 16, 16, 17, 17, 17, 16]
_ACTIVAS_4S = [40, 51, 62, 74, 85, 97, 108, 119, 130, 141, 151, 159]
_REVENUE_4S = [1500, 2168, 3100, 3700, 4250, 4850, 5400, 5950, 6500, 7050, 7550, 7950]


@dataclass(frozen=True)
class TeamTier:
    key: str
    label: str
    sales_count: int
    sales_base: int
    founder: int
    co_ceo: int
    dev: int
    dev_label: str
    cs_cm: int
    marketing_lead: int
    google_ads: int
    coordinador: int
    asesor: int
    contador: int
    ia: int
    hq: int
    servicios: int
    hosting: int
    meta_hi: int
    meta_lo: int
    valla_small: int
    valla_med: int
    valla_start: int
    reserva: int
    one_shots_extra: int
    safe_cap_base: int
    sales_curve_scale: float
    revenue_scale: float
    sales_commission_per_signing: int
    salary_note: str


TIERS: dict[str, TeamTier] = {
    "lean": TeamTier(
        key="lean",
        label="Lean — bootstrap 4× Sales (Carabobo)",
        sales_count=4,
        sales_base=350,
        founder=1000,
        co_ceo=1000,
        dev=600,
        dev_label="Dev junior (Flutter/Laravel)",
        cs_cm=400,
        marketing_lead=0,
        google_ads=0,
        coordinador=0,
        asesor=0,
        contador=330,
        ia=425,
        hq=500,
        servicios=80,
        hosting=154,
        meta_hi=800,
        meta_lo=500,
        valla_small=350,
        valla_med=0,
        valla_start=3,
        reserva=490,
        one_shots_extra=0,
        safe_cap_base=600_000,
        sales_curve_scale=1.0,
        revenue_scale=1.0,
        sales_commission_per_signing=40,
        salary_note="4× Sales bootstrap (350 c/u) — sueldos bajo AVGH; validar contratos [PENDIENTE]",
    ),
    "lean_plus": TeamTier(
        key="lean_plus",
        label="Lean+ — 4× Sales curva ×1,15 (mismo capital Lean)",
        sales_count=4,
        sales_base=350,
        founder=1000,
        co_ceo=1000,
        dev=600,
        dev_label="Dev junior (Flutter/Laravel)",
        cs_cm=400,
        marketing_lead=0,
        google_ads=0,
        coordinador=0,
        asesor=0,
        contador=330,
        ia=425,
        hq=500,
        servicios=80,
        hosting=154,
        meta_hi=800,
        meta_lo=500,
        valla_small=350,
        valla_med=0,
        valla_start=3,
        reserva=490,
        one_shots_extra=0,
        safe_cap_base=600_000,
        sales_curve_scale=1.15,
        revenue_scale=1.0,
        sales_commission_per_signing=40,
        salary_note="Mismo burn Lean; curva comercial +15% (~214 firmas Y1) — BE operativo ~M10",
    ),
    "base": TeamTier(
        key="base",
        label="Base — piso de mercado",
        sales_count=4,
        sales_base=400,
        founder=1200,
        co_ceo=1200,
        dev=1000,
        dev_label="Dev mid (Flutter/Laravel)",
        cs_cm=500,
        marketing_lead=300,
        google_ads=300,
        coordinador=350,
        asesor=120,
        contador=450,
        ia=425,
        hq=500,
        servicios=80,
        hosting=154,
        meta_hi=800,
        meta_lo=800,
        valla_small=0,
        valla_med=700,
        valla_start=2,
        reserva=10_590,
        one_shots_extra=900 + VIDEO_B2B_CASE,
        safe_cap_base=650_000,
        sales_curve_scale=1.0,
        revenue_scale=1.0,
        sales_commission_per_signing=45,
        salary_note="Piso mercado VE 2026 (AVGH profesional ~785, dev mid ~1.000)",
    ),
    "growth": TeamTier(
        key="growth",
        label="Growth — mercado pleno",
        sales_count=4,
        sales_base=450,
        founder=1500,
        co_ceo=1500,
        dev=1800,
        dev_label="Dev senior (Flutter/Laravel)",
        cs_cm=600,
        marketing_lead=150,
        google_ads=300,
        coordinador=450,
        asesor=120,
        contador=600,
        ia=425,
        hq=500,
        servicios=80,
        hosting=154,
        meta_hi=800,
        meta_lo=800,
        valla_small=0,
        valla_med=700,
        valla_start=2,
        reserva=20_000,
        one_shots_extra=900,
        safe_cap_base=720_000,
        sales_curve_scale=1.0,
        revenue_scale=1.0,
        sales_commission_per_signing=50,
        salary_note="Mercado pleno VE 2026 — comisiones Sales aparte del base",
    ),
    "blitz": TeamTier(
        key="blitz",
        label="Blitz — Carabobo 6× Sales",
        sales_count=6,
        sales_base=400,
        founder=1200,
        co_ceo=1200,
        dev=1000,
        dev_label="Dev mid (Flutter/Laravel)",
        cs_cm=550,
        marketing_lead=300,
        google_ads=500,
        coordinador=400,
        asesor=120,
        contador=450,
        ia=425,
        hq=500,
        servicios=80,
        hosting=154,
        meta_hi=800,
        meta_lo=800,
        valla_small=0,
        valla_med=700,
        valla_start=2,
        reserva=25_000,
        one_shots_extra=900 + VIDEO_B2B_CASE,
        safe_cap_base=780_000,
        sales_curve_scale=1.5,
        revenue_scale=1.0,
        sales_commission_per_signing=45,
        salary_note="6× Sales @ 400 — curva ×1,5 (captura >80% independientes Carabobo)",
    ),
}

ACTIVE_TIER = TIERS["lean"]


def _contingencia(m: int) -> int:
    return 1322 if m <= 2 else 1308 if m <= 6 else 1241


def _valla_for_month(t: TeamTier, m: int) -> int:
    if t.valla_med and m >= t.valla_start:
        return t.valla_med
    if t.valla_small and m >= t.valla_start:
        return t.valla_small
    return 0


def _meta_for_month(t: TeamTier, m: int) -> int:
    return t.meta_hi if m <= 6 else t.meta_lo


def fixed_payroll(t: TeamTier) -> int:
    return (
        t.founder
        + t.co_ceo
        + t.dev
        + t.sales_count * t.sales_base
        + t.cs_cm
        + t.marketing_lead
        + t.google_ads
        + t.coordinador
        + t.asesor
        + t.contador
        + t.ia
        + t.hq
        + t.servicios
        + t.hosting
    )


def burn_for_tier(t: TeamTier) -> list[int]:
    fixed = fixed_payroll(t)
    return [
        fixed + _meta_for_month(t, m) + _valla_for_month(t, m) + _contingencia(m)
        for m in MONTHS
    ]


def one_shots_for_tier(t: TeamTier) -> int:
    return LEGAL_SUBTOTAL + INTRO_LEAN + HQ_CAPEX_LEAN + t.one_shots_extra


def fase0_for_tier(t: TeamTier, burns: list[int]) -> tuple[int, int, int]:
    one = one_shots_for_tier(t)
    partial = round(burns[0] * 0.45)
    f0a = one + partial
    f0b = burns[0]
    f0c = burns[0]
    return f0a, f0b, f0c


def capital_for_tier(t: TeamTier, burns: list[int]) -> int:
    return one_shots_for_tier(t) + sum(burns) + t.reserva


def safe_cap_for_tier(t: TeamTier, capital: int, lean_capital: int) -> int:
    return round(t.safe_cap_base * capital / lean_capital)


def scale_curve(values: list[int], scale: float) -> list[int]:
    return [max(1, round(v * scale)) if v > 0 else 0 for v in values]


def revenue_curve_for_tier(t: TeamTier) -> list[int]:
    scaled = scale_curve(_REVENUE_4S, t.sales_curve_scale)
    return [round(r * t.revenue_scale) for r in scaled]


def first_be_month(revenue: list[int], burn: list[int]) -> int | None:
    for i, (rev, b) in enumerate(zip(revenue, burn)):
        if rev >= b:
            return i + 1
    return None


# --- Lean (Año 1 ancla) — calculado bottom-up ---
BURN = burn_for_tier(ACTIVE_TIER)
CONTINGENCIA_BY_MONTH = [_contingencia(m) for m in MONTHS]
META_BY_MONTH = [_meta_for_month(ACTIVE_TIER, m) for m in MONTHS]
VALLA_BY_MONTH = [_valla_for_month(ACTIVE_TIER, m) for m in MONTHS]

FIRMADAS = scale_curve(_FIRMADAS_4S, ACTIVE_TIER.sales_curve_scale)
ACTIVAS = scale_curve(_ACTIVAS_4S, ACTIVE_TIER.sales_curve_scale)
REVENUE = revenue_curve_for_tier(ACTIVE_TIER)

ONE_SHOTS_LEAN = one_shots_for_tier(ACTIVE_TIER)
FASE0_0A, FASE0_0B, FASE0_0C = fase0_for_tier(ACTIVE_TIER, BURN)
FASE0_TOTAL = FASE0_0A + FASE0_0B + FASE0_0C
RESERVA_LEAN = ACTIVE_TIER.reserva
SAFE_LEAN = capital_for_tier(ACTIVE_TIER, BURN)
CAJA_DAYD = SAFE_LEAN - FASE0_TOTAL
CAP_LEAN = safe_cap_for_tier(ACTIVE_TIER, SAFE_LEAN, SAFE_LEAN)

BURN_Y1 = sum(BURN)
REVENUE_Y1 = sum(REVENUE)
FCF_Y1 = REVENUE_Y1 - BURN_Y1
CAJA_M12_THEORETICAL = CAJA_DAYD + FCF_Y1
EQUILIBRIO_ACTIVAS_LEAN = ACTIVAS[-1]
EQUITY_INV = round(SAFE_LEAN / CAP_LEAN, 4)
MODEL_VERSION = "v3.8.2"
# Filas ESTA: anclas dinámicas en DETALLADO_ROWS (esta_burn_anchor_row, esta_sim_anchor_row)
ESTA_ROW_GAP = 1  # filas en blanco entre secciones apiladas
DET_SHEET = "Detallado de la inversión."
# PROYECCION §7 — anclas escenarios (no scale_curve 0.5 que invertía P10/P90)
P10_ACTIVAS_M12 = 120
P90_ACTIVAS_M12 = 200
P10_CASH_M12 = 25_000
P90_CASH_M12 = 52_000


def _equity_pct_es() -> str:
    return f"{EQUITY_INV * 100:.2f}".replace(".", ",")


def _founders_pct_es() -> str:
    return f"{(1 - EQUITY_INV) * 100:.2f}".replace(".", ",")


SALES_COMMISSION_M_AVG = round(
    sum(FIRMADAS) * ACTIVE_TIER.sales_commission_per_signing / 12
)

TIER_SUMMARY = {}
for key, tier in TIERS.items():
    burns = burn_for_tier(tier)
    cap = capital_for_tier(tier, burns)
    rev = revenue_curve_for_tier(tier)
    TIER_SUMMARY[key] = {
        "capital": cap,
        "burn_avg": round(sum(burns) / 12),
        "burn_y1": sum(burns),
        "burn_m12": burns[-1],
        "one_shots": one_shots_for_tier(tier),
        "fase0": sum(fase0_for_tier(tier, burns)),
        "caja_dayd": cap - sum(fase0_for_tier(tier, burns)),
        "safe_cap": safe_cap_for_tier(tier, cap, SAFE_LEAN),
        "activas_eq": scale_curve(_ACTIVAS_4S, tier.sales_curve_scale)[-1],
        "revenue_y1": sum(rev),
        "revenue_m12": rev[-1],
        "be_month": first_be_month(rev, burns),
        "caja_m12": cap - sum(fase0_for_tier(tier, burns)) + sum(rev) - sum(burns),
    }

# Filas ESTA SI VALE columna I (USD/mes ref.) — poblado en build_esta_si_vale
ESTA_BURN_ROWS: dict[str, int] = {}


def expense_lines_for_tier(t: TeamTier) -> list[tuple[str, str | None]]:
    sales_label = f"{t.sales_count}× Sales B2B"
    lines: list[tuple[str, str | None]] = [
        (t.dev_label, t.dev_label),
        (sales_label, sales_label),
        ("CS + Community Manager", "CS + Community Manager"),
    ]
    if t.marketing_lead:
        lines.append(("Marketing Lead / diseñador B2B", "Marketing Lead / diseñador B2B"))
    if t.google_ads:
        lines.append(("Google Ads / YouTube B2B", "Google Ads / YouTube B2B"))
    if t.coordinador:
        lines.append(("Coordinador Partners Logísticos", "Coordinador Partners Logísticos"))
    if t.asesor:
        lines.append(("Asesor regulatorio farmacéutico", "Asesor regulatorio farmacéutico"))
    lines.extend([
        ("Founder + Co-CEO", "Founder + Co-CEO"),
        ("Herramientas IA", "Herramientas IA"),
        ("Contador + Abogado", "Contador + Abogado"),
        ("HQ casa (arriendo + servicios)", "HQ casa (arriendo + servicios)"),
        ("Servicios oficina (electricidad, agua, internet)", "Servicios oficina (electricidad, agua, internet)"),
        ("Hosting + SaaS + Firebase OTP", "Hosting + SaaS + Firebase OTP"),
        ("Meta Ads", None),
    ])
    if t.google_ads:
        lines.append(("Google Ads / YouTube B2B (farmacia)", None))
    if t.valla_med:
        lines.append(("Valla mediana (desde M2)", None))
    else:
        lines.append(("Valla pequeña (desde M3)", None))
    lines.append(("Material + transporte + contingencia (incl. buffer VE 20%)", None))
    return lines


EXPENSE_LINES = expense_lines_for_tier(ACTIVE_TIER)


def burn_catalog_for_tier(
    t: TeamTier,
    meta_avg: int,
    valla_avg: int,
    cont_avg: int,
) -> list[tuple[str, int, str]]:
    """Catálogo burn mensual (USD/mes) compartido ESTA + Detallado."""
    catalog: list[tuple[str, int, str]] = [
        (t.dev_label, t.dev, "Nómina v3"),
        (f"{t.sales_count}× Sales B2B (base)", t.sales_count * t.sales_base, "PRESUPUESTO §2.3"),
        ("CS + Community Manager", t.cs_cm, ""),
        ("Founder + Co-CEO", t.founder + t.co_ceo, ""),
        ("Herramientas IA", t.ia, ""),
        ("Contador + Abogado", t.contador, ""),
        ("HQ casa (arriendo + servicios)", t.hq, ""),
        ("Servicios oficina", t.servicios, "§2.3.1"),
        ("Hosting + SaaS + Firebase OTP", t.hosting, "§2.3.1"),
        ("Meta Ads (prom.)", meta_avg, "tramos"),
    ]
    if t.google_ads:
        catalog.append(
            ("Google Ads / YouTube B2B (farmacia)", t.google_ads, "v3.2 Base+ — filtrado Pizza")
        )
    if t.marketing_lead:
        catalog.insert(3, ("Marketing Lead / diseñador B2B", t.marketing_lead, "Base/Growth"))
    if t.valla_med:
        catalog.append(("Valla mediana (desde M2)", valla_avg, ""))
    else:
        catalog.append(("Valla pequeña (desde M3)", valla_avg, ""))
    if t.coordinador:
        insert_at = 4 if t.marketing_lead else 3
        catalog.insert(insert_at, ("Coordinador Partners Logísticos", t.coordinador, ""))
    if t.asesor:
        catalog.insert(5, ("Asesor regulatorio farmacéutico", t.asesor, ""))
    catalog.append(
        ("Material + transporte + contingencia (incl. buffer VE 20%)", cont_avg, "PRESUPUESTO §6")
    )
    return catalog


def mo_operativa_monthly_lines(t: TeamTier) -> list[tuple[str, int]]:
    """MO operativa post-Day-D — USD/mes por rol (espejo burn fijo)."""
    lines: list[tuple[str, int]] = [
        ("Founder CEO/CTO", t.founder),
        ("Co-CEO Head Sales&Ops", t.co_ceo),
        (t.dev_label.split("(")[0].strip(), t.dev),
        (f"{t.sales_count}× Sales B2B (base c/u)", t.sales_count * t.sales_base),
        ("CS + Community Manager", t.cs_cm),
    ]
    if t.marketing_lead:
        lines.append(("Marketing Lead / diseñador B2B", t.marketing_lead))
    if t.google_ads:
        lines.append(("Google Ads / YouTube B2B (farmacia)", t.google_ads))
    if t.coordinador:
        lines.append(("Coordinador Partners Logísticos", t.coordinador))
    if t.asesor:
        lines.append(("Asesor regulatorio farmacéutico", t.asesor))
    lines.extend([
        ("Herramientas IA", t.ia),
        ("Contador + Abogado", t.contador),
    ])
    return lines


def _lateral_g_formula(row: int, *, link_detallado: bool) -> str:
    if link_detallado:
        return f"='{DET_SHEET}'!G{row}"
    return f"=G{row}"


def _build_resumen_lateral_km(
    ws,
    start_row: int,
    *,
    include_timeline: bool = False,
    link_detallado: bool = False,
) -> None:
    """Caja resumen cols K–M — espejo Detallado; Hoja3 añade sub-fases 0a–0c."""
    col_l, col_m = 11, 13
    dr = DETALLADO_ROWS
    g = lambda row: _lateral_g_formula(row, link_detallado=link_detallado)

    ws.merge_cells(start_row=start_row, start_column=col_l, end_row=start_row, end_column=col_m)
    title = ws.cell(start_row, col_l)
    if include_timeline:
        title.value = f"TIMELINE + RESUMEN FASE 0 ({MODEL_VERSION})"
    else:
        title.value = f"RESUMEN INVERSIÓN LEAN ({MODEL_VERSION})"
    title.font = Font(bold=True, size=10, color="FFFFFF")
    title.fill = HDR_FILL
    title.alignment = Alignment(horizontal="center")

    rows: list[tuple[str, str | None]] = [
        ("One-shots (legal+intro+HQ)", g(dr["one_shots"])),
        ("Fase 0 operativa (sin one-shots)", g(dr["cross_ops_fase0"])),
        ("Fase 0 total (T+0→Day-D)", g(dr["subtotal_fase0"])),
    ]
    if include_timeline:
        rows.extend([
            ("Sub-fase 0a (T+0→T+30)", g(dr["fase0_0a_row"])),
            ("Sub-fase 0b (T+30→T+60)", g(dr["fase0_0b_row"])),
            ("Sub-fase 0c (T+60→T+90)", g(dr["fase0_0c_row"])),
        ])
    rows.extend([
        ("Caja al Day-D (inicio M1)", g(dr["caja_dayd"])),
        ("Burn 12 meses (M1–M12)", g(dr["burn_row"])),
        ("Reserva caja", g(dr["reserva_row"])),
        ("TOTAL capital SAFE", g(dr["total_safe"])),
    ])

    rr = start_row + 1
    total_m_row: int | None = None
    for label, formula in rows:
        set_text(ws, rr, col_l, label)
        if formula:
            set_num(ws, rr, col_m, formula=formula)
            if "TOTAL" in label:
                ws.cell(rr, col_m).font = TOTAL_FONT
                total_m_row = rr
        rr += 1

    equity_pct = round(EQUITY_INV * 100, 2)
    set_text(ws, rr, col_l, f"SAFE cap ref. → ~{equity_pct}% equity")
    if total_m_row is not None:
        set_num(ws, rr, col_m, formula=f"=M{total_m_row}/{CAP_LEAN}", fmt=PCT_FMT)
    else:
        set_text(ws, rr, col_m, f"Cap {CAP_LEAN:,} USD".replace(",", "."))
    ws.cell(rr, col_m).font = REF_FONT
    rr += 2

    set_text(ws, rr, col_l, "Ref. mensual post-Day-D", True)
    ws.cell(rr, col_l).fill = SEC_FILL
    rr += 1
    set_text(ws, rr, col_l, "MO operativa (subtotal)")
    set_num(ws, rr, col_m, formula=g(dr["subtotal_mo_operativa"]))
    rr += 1
    set_text(ws, rr, col_l, "Marketing mensual (subtotal)")
    set_num(ws, rr, col_m, formula=g(dr["subtotal_mkt_mensual"]))
    rr += 1
    ws.merge_cells(start_row=rr, start_column=col_l, end_row=rr, end_column=col_m)
    note = ws.cell(rr, col_l)
    note.value = (
        "Incluido en burn M1–M12. No sumar a one-shots ni duplicar MO Fase 0. "
        "Unit economics → pestaña ESTA SI VALE."
    )
    note.font = REF_FONT
    note.fill = REF_FILL
    note.alignment = Alignment(wrap_text=True)

    ws.column_dimensions["K"].width = 28
    ws.column_dimensions["M"].width = 14


def _build_detallado_resumen_lateral(ws, **kwargs) -> None:
    """Wrapper — lateral Detallado sin timeline."""
    _build_resumen_lateral_km(ws, 10, include_timeline=False, link_detallado=False)

YEAR_DATA = [
    (1, ACTIVAS[-1], REVENUE_Y1, BURN_Y1, "Lean v3 bottom-up"),
    (2, 220, 102_000, 96_000, "PROYECCION §2"),
    (3, 440, 192_000, 108_000, "PROYECCION §3"),
    (4, 520, 240_000, 130_000, "SUPUESTO MODELO"),
    (5, 600, 300_000, 150_000, "SUPUESTO MODELO"),
]

# Pesos trimestrales revenue/costos años 2–3 (PROYECCION §2.2 / §3.2)
Y2_REV_Q = [6525, 6525, 6525, 7875, 7875, 7875, 9000, 9000, 9000, 9875, 9875, 9875]
Y2_COST_Q = [7750, 7750, 7750, 8000, 8000, 8000, 8250, 8250, 8250, 8250, 8250, 8250]
Y3_REV_Q = [11700, 11700, 11700, 13950, 13950, 13950, 16200, 16200, 16200, 18900, 18900, 18900]
Y3_COST_Q = [7500, 7500, 7500, 8500, 8500, 8500, 9500, 9500, 9500, 10500, 10500, 10500]

YEAR_ROWS: dict[int, dict[str, int]] = {}
DETALLADO_ROWS: dict[str, int] = {}

# Styles — Zonix brand
HDR_FILL = PatternFill("solid", fgColor="1E2A5A")
TEAL_FILL = PatternFill("solid", fgColor="0F4C5C")
SEC_FILL = PatternFill("solid", fgColor="E8EEF2")
REF_FILL = PatternFill("solid", fgColor="D9D9D9")
REF_FONT = Font(italic=True, size=9, color="444444")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=12, color="1E2A5A")
SUB_FONT = Font(bold=True, size=10)
TOTAL_FONT = Font(bold=True, size=10)
NUM_FMT = '#,##0.00;[Red](#,##0.00)'
PCT_FMT = "0.00%"
THIN = Side(style="thin", color="CCCCCC")
DOUBLE = Side(style="double", color="1E2A5A")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOTAL_BORDER = Border(left=THIN, right=THIN, top=DOUBLE, bottom=THIN)


def col_month(m: int) -> int:
    return 2 + m


def col_total() -> int:
    return 15


def style_header_row(ws, row: int, cols: range):
    for c in cols:
        cell = ws.cell(row, c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def style_section_row(ws, row: int, cols: range = range(2, 16)):
    for c in cols:
        cell = ws.cell(row, c)
        cell.fill = SEC_FILL
        cell.font = SUB_FONT
        cell.border = BORDER


def style_total_row(ws, row: int, cols: range = range(2, 16)):
    for c in cols:
        cell = ws.cell(row, c)
        cell.font = TOTAL_FONT
        cell.border = TOTAL_BORDER
        if c >= 3:
            cell.number_format = NUM_FMT


def set_num(ws, row: int, col: int, value=None, formula: str | None = None, fmt=NUM_FMT):
    cell = ws.cell(row, col)
    if formula:
        cell.value = formula
    elif value is not None:
        cell.value = value
    cell.number_format = fmt
    cell.border = BORDER


def set_text(ws, row: int, col: int, value: str, bold: bool = False):
    cell = ws.cell(row, col)
    cell.value = value
    cell.border = BORDER
    if bold:
        cell.font = SUB_FONT if not isinstance(bold, bool) or bold else SUB_FONT


def month_headers(ws, row: int):
    for m in MONTHS:
        set_text(ws, row, col_month(m), f"Mes {m}")
    set_text(ws, row, col_total(), "TOTAL", True)
    style_header_row(ws, row, range(3, 16))


def fill_month_row(ws, row: int, values: list | None = None, total_formula: str | None = None, formula_template: str | None = None):
    if values:
        for i, v in enumerate(values):
            if v is None:
                continue
            if isinstance(v, str) and v.startswith("="):
                set_num(ws, row, col_month(i + 1), formula=v)
            else:
                set_num(ws, row, col_month(i + 1), value=v)
    elif formula_template:
        for m in range(12):
            c = col_month(m + 1)
            cl = get_column_letter(c)
            set_num(ws, row, c, formula=formula_template.format(col=cl, row=row))
    tc = col_total()
    if total_formula:
        set_num(ws, row, tc, formula=total_formula)
    else:
        sc, ec = get_column_letter(col_month(1)), get_column_letter(col_month(12))
        set_num(ws, row, tc, formula=f"=SUM({sc}{row}:{ec}{row})")


def expense_value(line_name: str, month_idx: int) -> int:
    if line_name == "Meta Ads":
        return META_BY_MONTH[month_idx]
    if "Valla pequeña" in line_name or "Valla mediana" in line_name:
        return VALLA_BY_MONTH[month_idx]
    if "contingencia" in line_name.lower():
        return CONTINGENCIA_BY_MONTH[month_idx]
    return 0


def quarterly_monthly(year_num: int, kind: str) -> list[float]:
    if year_num == 2:
        base = Y2_REV_Q if kind == "rev" else Y2_COST_Q
    elif year_num == 3:
        base = Y3_REV_Q if kind == "rev" else Y3_COST_Q
    else:
        yd = YEAR_DATA[year_num - 1]
        val = (yd[2] if kind == "rev" else yd[3]) / 12
        return [val] * 12
    total = sum(base)
    target = YEAR_DATA[year_num - 1][2 if kind == "rev" else 3]
    factor = target / total if total else 1
    return [round(v * factor, 2) for v in base]


def _det_note(ws, r: int, text: str, ref: bool = False) -> int:
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
    set_text(ws, r, 3, text)
    if ref:
        ws.cell(r, 3).fill = REF_FILL
        ws.cell(r, 3).font = REF_FONT
    return r + 2


def _det_section_title(ws, r: int, title: str, *, ref: bool = False) -> int:
    set_text(ws, r, 3, title, True)
    fill = REF_FILL if ref else TEAL_FILL
    ws.cell(r, 3).fill = fill
    ws.cell(r, 3).font = (
        Font(bold=True, color="444444", size=10)
        if ref
        else Font(bold=True, color="FFFFFF", size=10)
    )
    return r + 1


def _det_items(
    ws,
    r: int,
    items: list[tuple[str, float, float]],
    *,
    ref: bool = False,
    subtotal_label: str | None = None,
    subtotal_key: str | None = None,
) -> tuple[int, int | None, int]:
    start = r
    for label, qty, price in items:
        set_text(ws, r, 3, label)
        set_num(ws, r, 5, value=qty)
        set_num(ws, r, 6, value=price)
        set_num(ws, r, 7, formula=f"=E{r}*F{r}")
        if ref:
            for col in (3, 5, 6, 7):
                ws.cell(r, col).fill = REF_FILL
        r += 1
    sub_row = None
    if subtotal_label:
        set_text(ws, r, 3, subtotal_label, True)
        set_num(ws, r, 7, formula=f"=SUM(G{start}:G{r-1})")
        if ref:
            ws.cell(r, 3).fill = REF_FILL
            ws.cell(r, 7).fill = REF_FILL
            set_text(ws, r, 9, "[EDITAR — no suma ancla]")
        sub_row = r
        if subtotal_key:
            DETALLADO_ROWS[subtotal_key] = r
        r += 2
    return r, sub_row, start


def build_detallado(ws):
    ws.title = "Detallado de la inversión."
    ws.merge_cells("C4:I4")
    c = ws["C4"]
    c.value = f"Zonix Pharma — Marketplace farmacéutico VE (piloto Valencia) — {MODEL_VERSION} bottom-up"
    c.font = TITLE_FONT

    headers = {
        8: ("Descripción", 3),
        9: ("Cant.", 5),
        10: ("USD/u", 6),
        11: ("Total USD", 7),
        12: ("% Inversión", 8),
        13: ("Notas / En ESTA", 9),
    }
    for row_n, (h, col) in headers.items():
        set_text(ws, row_n, col, h, True)
    style_header_row(ws, 8, [3, 5, 6, 7, 8, 9])

    ref_rows: set[int] = set()
    r = 10

    # --- Constitución empresa Zonix Pharma C.A. (costos sinceros, ref. trámites reales VE) ---
    r = _det_section_title(ws, r, "CONSTITUCIÓN DE LA EMPRESA (Zonix Pharma C.A.)")
    r, legal_sub, legal_start = _det_items(
        ws,
        r,
        [
            ("Registro Mercantil + acta constitutiva", 1, 1200),
            ("SENIAT — inscripción / trámites RIF", 1, 600),
            ("Licencia / permisos municipales HQ", 1, 1500),
            ("Permiso bomberos (local operativo)", 1, 1100),
            ("Apertura cuenta bancaria empresa", 1, 100),
            ("Sello digital + factura digital SENIAT", 1, 200),
            ("Registro marca SAPI (paralelo)", 1, 350),
        ],
        subtotal_label="Subtotal constitución empresa",
        subtotal_key="subtotal_legal",
    )
    DETALLADO_ROWS["legal_start_row"] = legal_start
    DETALLADO_ROWS["legal_end_row"] = legal_sub - 1 if legal_sub else legal_start
    r = _det_note(
        ws,
        r,
        "Costos legales de constitución C.A. + local HQ (referencia real VE). "
        "No incluye permisos de farmacia aliada ni operación de medicamentos. "
        "Pack documental conservador 1.150–1.450 en ESTRUCTURA_LEGAL §1.4 — validar con abogado [PENDIENTE].",
    )

    r = _det_note(
        ws,
        r,
        "MATERIA PRIMA: No aplica — ZonixPharma es servicio tecnológico (marketplace). "
        "Equivalente digital: SMS OTP, Pusher, Maps, hosting → fila «Hosting + SaaS + Firebase OTP» "
        "en ESTA SI VALE (~154 USD/mes) + contingencia en burn.",
    )

    # --- Capa A: intro Lean ---
    r = _det_section_title(ws, r, "INTRODUCCIÓN AL MERCADO / DEMO — Lean (ancla Fase 0)")
    r, intro_sub, intro_start = _det_items(
        ws,
        r,
        [
            ("Website landing + dominio", 1, 800),
            ("Brochures y material impreso B2B", 1, 400),
            ("Tablets demo farmacia", 2, 350),
            ("Visitas comerciales pre-Day-D", 1, 500),
            ("Reservas operativas primer mes", 1, 608),
            ("Video pitch B2B farmacia (one-shot)", 1, VIDEO_B2B_INTRO),
        ],
        subtotal_label="Subtotal intro mercado Lean",
        subtotal_key="subtotal_intro",
    )
    DETALLADO_ROWS["intro_start_row"] = intro_start
    DETALLADO_ROWS["intro_end_row"] = intro_sub - 1 if intro_sub else intro_start

    # --- Marketing pre Lean ---
    r = _det_section_title(ws, r, "MARKETING PRE-LANZAMIENTO — Lean (3 meses, ancla)")
    r, mkt_pre_sub, mkt_pre_start = _det_items(
        ws,
        r,
        [
            ("Meta Ads pre-Day-D", 3, 400),
            ("Material POP / señalética", 1, 600),
            ("Demos B2B adicionales", 1, 450),
        ],
        subtotal_label="Subtotal marketing pre-lanzamiento Lean",
        subtotal_key="subtotal_mkt_pre",
    )
    DETALLADO_ROWS["mkt_pre_start_row"] = mkt_pre_start
    DETALLADO_ROWS["mkt_pre_end_row"] = mkt_pre_sub - 1 if mkt_pre_sub else mkt_pre_start

    # --- Mano de obra Fase 0 (desglose Lean v3 — incluye Dev en nómina) ---
    t = ACTIVE_TIER
    r = _det_section_title(
        ws,
        r,
        "MANO DE OBRA FASE 0 (3 meses — incluida en Fase 0, no duplicar en one-shots)",
    )
    r = _det_note(
        ws,
        r,
        f"{t.salary_note}. Tarifas mensuales × 3 meses pre-Day-D. "
        f"Comisiones Sales: ${t.sales_commission_per_signing}/firma (estimado en ESTA, no duplicado aquí).",
    )
    mo_items = [
        (f"Founder CEO/CTO ({t.founder}/mes)", 3, t.founder),
        (f"Co-CEO Head Sales&Ops ({t.co_ceo}/mes)", 3, t.co_ceo),
        (f"{t.dev_label} ({t.dev}/mes)", 3, t.dev),
        (f"{t.sales_count}× Sales B2B ({t.sales_base}/mes c/u base)", 3, t.sales_count * t.sales_base),
        (f"CS + Community Manager ({t.cs_cm}/mes)", 3, t.cs_cm),
        (f"Herramientas IA ({t.ia}/mes)", 3, t.ia),
        (f"Contador + Abogado ({t.contador}/mes)", 3, t.contador),
    ]
    if t.coordinador:
        mo_items.insert(-2, (f"Coordinador Partners ({t.coordinador}/mes)", 3, t.coordinador))
    if t.asesor:
        mo_items.insert(-2, (f"Asesor regulatorio ({t.asesor}/mes)", 3, t.asesor))
    r, mo_f0_sub, mo_f0_start = _det_items(
        ws,
        r,
        mo_items,
        subtotal_label="Subtotal mano de obra Fase 0 (3 meses)",
        subtotal_key="subtotal_mo_fase0",
    )
    DETALLADO_ROWS["mo_fase0_start_row"] = mo_f0_start
    DETALLADO_ROWS["mo_fase0_end_row"] = mo_f0_sub - 1 if mo_f0_sub else mo_f0_start

    # --- Capa B: transporte / movilidad B2B ---
    r = _det_section_title(ws, r, "TRANSPORTE Y MOVILIDAD B2B (referencia desglose)")
    r = _det_note(
        ws,
        r,
        "Ancla Lean: ~500 USD en «Visitas comerciales» (intro) + ~280/mes en burn "
        "«Material + transporte + contingencia». Bloque explicativo — no suma ancla.",
    )
    r, transport_sub, transport_start = _det_items(
        ws,
        r,
        [
            ("Visitas comerciales Valencia (combustible / terceros)", 3, 167),
            ("Logística material demo (tablets, POP)", 1, 200),
            ("Gira medios / eventos farmacia (opcional Lean)", 1, 350),
        ],
        subtotal_label="Subtotal transporte referencia Fase 0",
        subtotal_key="subtotal_transporte_ref",
    )
    DETALLADO_ROWS["transport_start_row"] = transport_start
    DETALLADO_ROWS["transport_end_row"] = transport_sub - 1 if transport_sub else transport_start

    # --- Capa A: HQ + Fase 0 ---
    r = _det_section_title(ws, r, "HQ Y CAPEX — Lean (ancla one-shots)")
    r, hq_sub, hq_start = _det_items(
        ws,
        r,
        [
            ("Depósito + amueblado HQ casa", 1, 1250),
            ("Adecuación HQ (pintura, señalética)", 1, 400),
            ("PC recepción / ventas", 1, 800),
            ("PC administración", 1, 850),
            ("PC desarrollo", 1, 950),
            ("PC puesto flexible", 1, 1100),
        ],
        subtotal_label="Subtotal HQ y CapEx",
        subtotal_key="subtotal_hq",
    )
    # Solo las 4 PCs (filas inmediatamente antes del subtotal HQ)
    DETALLADO_ROWS["pc_start_row"] = hq_sub - 4
    DETALLADO_ROWS["pc_end_row"] = hq_sub - 1
    DETALLADO_ROWS["hq_start_row"] = hq_start
    DETALLADO_ROWS["hq_end_row"] = hq_sub - 1 if hq_sub else hq_start
    DETALLADO_ROWS["hq_adecuacion_end_row"] = hq_start + 1

    r = _det_section_title(ws, r, "FASE 0 — OUTFLOW OPERATIVO (T+0→Day-D, bottom-up)")
    r, fase0_sub, fase0_items_start = _det_items(
        ws,
        r,
        [
            ("Sub-fase 0a (T+0 a T+30)", 1, FASE0_0A),
            ("Sub-fase 0b (T+30 a T+60)", 1, FASE0_0B),
            ("Sub-fase 0c (T+60 a T+90)", 1, FASE0_0C),
        ],
        subtotal_label="Subtotal Fase 0",
        subtotal_key="subtotal_fase0",
    )
    DETALLADO_ROWS["fase0_0a_row"] = fase0_items_start
    DETALLADO_ROWS["fase0_0b_row"] = fase0_items_start + 1
    DETALLADO_ROWS["fase0_0c_row"] = fase0_items_start + 2

    set_num(ws, r, 7, formula=f"={SAFE_LEAN}-G{DETALLADO_ROWS['subtotal_fase0']}")
    set_text(ws, r, 3, f"Caja al Day-D ({SAFE_LEAN:,} − Fase 0)".replace(",", "."), True)
    DETALLADO_ROWS["caja_dayd"] = r
    r += 2

    one_shot_row = r
    set_text(ws, r, 3, "TOTAL one-shots Lean (legal+intro+HQ)", True)
    set_num(
        ws,
        r,
        7,
        formula=(
            f"=G{DETALLADO_ROWS['subtotal_legal']}+G{DETALLADO_ROWS['subtotal_intro']}"
            f"+G{DETALLADO_ROWS['subtotal_hq']}"
        ),
    )
    DETALLADO_ROWS["one_shots"] = r
    r += 2

    cross_row = r
    set_text(ws, r, 3, "Validación: Fase 0 − one-shots − mkt pre (ops burn ~15.999)", True)
    set_num(
        ws,
        r,
        7,
        formula=(
            f"=G{DETALLADO_ROWS['subtotal_fase0']}-G{one_shot_row}"
            f"-G{DETALLADO_ROWS['subtotal_mkt_pre']}"
        ),
    )
    set_text(ws, r, 9, "≈ delta vs subtotal MO + HQ/servicios en Fase 0")
    DETALLADO_ROWS["cross_ops_fase0"] = r
    r += 2

    # --- MO operativa post-Day-D (Lean) ---
    r = _det_section_title(
        ws,
        r,
        "MANO DE OBRA OPERATIVA (post-Day-D — USD/mes, incluida en burn M1–M12)",
    )
    r = _det_note(
        ws,
        r,
        "Espejo nómina recurrente ESTA SI VALE. No sumar a one-shots ni a Fase 0.",
    )
    mo_op_start = r
    for label, usd_m in mo_operativa_monthly_lines(t):
        set_text(ws, r, 3, label)
        set_num(ws, r, 6, value=usd_m)
        set_text(ws, r, 8, "mensual")
        set_text(ws, r, 9, "→ burn Año 1")
        r += 1
    set_text(ws, r, 3, "Subtotal MO operativa (fija/mes)", True)
    set_num(ws, r, 7, formula=f"=SUM(F{mo_op_start}:F{r-1})")
    DETALLADO_ROWS["subtotal_mo_operativa"] = r
    mo_subtotal_row = r
    r += 2

    # --- Marketing mensual recurrente (montos, no solo ref ESTA) ---
    meta_avg = round(sum(META_BY_MONTH) / 12)
    valla_avg = round(sum(VALLA_BY_MONTH) / 12)
    cont_avg = round(sum(CONTINGENCIA_BY_MONTH) / 12)
    r = _det_section_title(
        ws,
        r,
        "MARKETING Y OPEX MENSUAL RECURRENTE (ref. ESTA → Año 1)",
    )
    r = _det_note(
        ws,
        r,
        "Incluido en burn M1–M12. No sumar a one-shots ni duplicar MO Fase 0.",
    )
    mkt_start = r
    mkt_fixed_end = r
    for label, val, src in burn_catalog_for_tier(t, meta_avg, valla_avg, cont_avg):
        if label in (
            t.dev_label,
            f"{t.sales_count}× Sales B2B (base)",
            "CS + Community Manager",
            "Founder + Co-CEO",
            "Herramientas IA",
            "Contador + Abogado",
        ):
            continue
        if t.coordinador and "Coordinador" in label:
            continue
        if t.asesor and "Asesor" in label:
            continue
        if t.marketing_lead and "Marketing Lead" in label:
            continue
        set_text(ws, r, 3, label)
        set_num(ws, r, 6, value=val)
        set_num(ws, r, 7, formula=f"=F{r}")
        set_text(ws, r, 8, "mensual")
        note = f"→ ESTA: {src}" if src else "→ ESTA / Año 1"
        if "HQ casa" in label:
            DETALLADO_ROWS["hq_alquiler_row"] = r
            note = "Alquiler HQ post-Day-D — incluido en burn M1–M12"
        if "Hosting" in label:
            DETALLADO_ROWS["hosting_saas_row"] = r
            note = f"SaaS ~{HOSTING_SAAS_M} + SMS ~{HOSTING_SMS_M} (§2.3.1)"
        set_text(ws, r, 9, note)
        if "Hosting" in label:
            r += 1
            set_text(ws, r, 3, "  └ Hosting cloud / SaaS (desglose)")
            set_text(ws, r, 9, f"~{HOSTING_SAAS_M} USD/mes — footnote, no sumar")
            ws.cell(r, 3).font = REF_FONT
            ws.cell(r, 9).font = REF_FONT
            r += 1
            set_text(ws, r, 3, "  └ Firebase SMS / OTP (pico conservador)")
            set_text(ws, r, 9, f"~{HOSTING_SMS_M} USD/mes — footnote, no sumar")
            ws.cell(r, 3).font = REF_FONT
            ws.cell(r, 9).font = REF_FONT
        mkt_fixed_end = r
        r += 1
    set_text(ws, r, 3, "Subtotal marketing + opex variable/mes", True)
    set_num(ws, r, 7, formula=f"=SUM(F{mkt_start}:F{mkt_fixed_end})")
    DETALLADO_ROWS["subtotal_mkt_mensual"] = r
    mkt_subtotal_row = r
    r += 1
    set_text(ws, r, 3, f"Sales comisiones (estim. prom/mes — ${t.sales_commission_per_signing}/firma)")
    set_num(ws, r, 6, value=SALES_COMMISSION_M_AVG)
    set_text(ws, r, 8, "mensual")
    set_text(ws, r, 9, "Variable — no incluido en burn fijo Lean")
    ws.cell(r, 3).fill = REF_FILL
    ws.cell(r, 3).font = REF_FONT
    r += 2

    set_text(ws, r, 3, "USE OF FUNDS 12M (post Day-D) — bottom-up", True)
    r += 1
    burn_row = r
    set_text(ws, r, 3, "Burn M1-M12")
    set_num(ws, r, 7, value=BURN_Y1)
    DETALLADO_ROWS["burn_row"] = r
    r += 1
    reserva_row = r
    set_text(ws, r, 3, "Reserva caja Lean")
    set_num(ws, r, 7, value=RESERVA_LEAN)
    DETALLADO_ROWS["reserva_row"] = r
    r += 1
    total_row = r
    set_text(ws, r, 3, "TOTAL capital SAFE Lean", True)
    set_num(ws, r, 7, formula=f"=G{one_shot_row}+G{burn_row}+G{reserva_row}")
    DETALLADO_ROWS["total_safe"] = r

    _build_detallado_resumen_lateral(ws)

    for rr in range(10, r + 1):
        if rr in ref_rows:
            continue
        gval = ws.cell(rr, 7).value
        if gval is not None:
            set_num(ws, rr, 8, formula=f"=IF(G{rr}=0,\"\",G{rr}/G{total_row})", fmt=PCT_FMT)

    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 28


def build_hoja1(ws):
    """Hoja1 — 8 bloques Pizza con datos ZonixPharma (espejo Detallado) + panel refs."""
    ws.title = "Hoja1"
    dr = DETALLADO_ROWS

    set_text(ws, 1, 5, f"Zonix Pharma — inversión Fase 0 (espejo Detallado {MODEL_VERSION})", True)
    ws["E1"].font = TITLE_FONT

    ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=9)
    disclaimer = ws.cell(3, 5)
    disclaimer.value = (
        "Datos ZonixPharma — enlaces a Detallado. One-shots en Hoja3 Sección A. "
        "No sumar bloques al TOTAL SAFE. Marketplace SaaS (sin rubros producción física)."
    )
    disclaimer.font = REF_FONT
    disclaimer.fill = REF_FILL
    disclaimer.alignment = Alignment(wrap_text=True)

    set_text(ws, 5, 5, "Descripción", True)
    set_text(ws, 5, 6, "Cant.", True)
    set_text(ws, 5, 7, "USD/u", True)
    set_text(ws, 5, 8, "Total USD", True)
    set_text(ws, 5, 9, "% SAFE", True)
    style_header_row(ws, 5, [5, 6, 7, 8, 9])

    _build_hoja1_zonix_blocks(ws, 6, dr)
    _build_hoja1_panel_jk(ws, dr)

    ws.column_dimensions["E"].width = 44
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 28
    ws.column_dimensions["K"].width = 14


def build_hoja2(ws):
    ws.title = "Hoja2"
    set_text(ws, 1, 3, "mensual", True)
    refs = [
        ("Meta Ads M1-6 (USD)", 800),
        ("Meta Ads M7-12 (USD)", 500),
        ("Valla pequeña desde M3 (USD)", 350),
    ]
    for i, (label, val) in enumerate(refs, start=2):
        set_text(ws, i, 2, label)
        set_num(ws, i, 3, value=val)


def _mirror_det_row(
    ws,
    row: int,
    det_row: int,
    start_col: int = 5,
    pct_col: int | None = None,
) -> None:
    """Espejo Cant × USD/u → Total desde Detallado (sin duplicar montos)."""
    sc = start_col
    ws.cell(row, sc).value = f"='{DET_SHEET}'!C{det_row}"
    ws.cell(row, sc + 1).value = f"='{DET_SHEET}'!E{det_row}"
    ws.cell(row, sc + 2).value = f"='{DET_SHEET}'!F{det_row}"
    ws.cell(row, sc + 3).value = f"='{DET_SHEET}'!G{det_row}"
    for col in range(sc, sc + 4):
        ws.cell(row, col).border = BORDER
    if pct_col is not None:
        set_num(ws, row, pct_col, formula=f"='{DET_SHEET}'!H{det_row}", fmt=PCT_FMT)
        ws.cell(row, pct_col).border = BORDER


def _mirror_detallado_block(
    ws,
    row: int,
    title: str,
    start_row: int,
    end_row: int,
    subtotal_key: str | None = None,
    *,
    start_col: int = 5,
    ref: bool = False,
    pct_col: int | None = None,
) -> int:
    """Espejo read-only Detallado — compartido Hoja1 y ESTA SI VALE."""
    return _mirror_block(
        ws,
        row,
        title,
        start_row,
        end_row,
        subtotal_key,
        start_col=start_col,
        ref=ref,
        pct_col=pct_col,
    )


def _mirror_block(
    ws,
    row: int,
    title: str,
    start_row: int,
    end_row: int,
    subtotal_key: str | None = None,
    *,
    start_col: int = 5,
    ref: bool = False,
    pct_col: int | None = None,
) -> int:
    sc = start_col
    total_col = sc + 3
    set_text(ws, row, sc, title, True)
    ws.cell(row, sc).fill = REF_FILL if ref else TEAL_FILL
    ws.cell(row, sc).font = Font(
        bold=True,
        color="444444" if ref else "FFFFFF",
        size=10,
    )
    row += 1
    for det_row in range(start_row, end_row + 1):
        _mirror_det_row(ws, row, det_row, start_col, pct_col=pct_col)
        if ref:
            fill_cols = list(range(sc, sc + 4))
            if pct_col is not None:
                fill_cols.append(pct_col)
            for col in fill_cols:
                ws.cell(row, col).fill = REF_FILL
        row += 1
    if subtotal_key and subtotal_key in DETALLADO_ROWS:
        sub_r = DETALLADO_ROWS[subtotal_key]
        ws.cell(row, sc).value = f"='{DET_SHEET}'!C{sub_r}"
        ws.cell(row, sc).font = SUB_FONT
        set_num(ws, row, total_col, formula=f"='{DET_SHEET}'!G{sub_r}")
        if pct_col is not None:
            set_num(ws, row, pct_col, formula=f"='{DET_SHEET}'!H{sub_r}", fmt=PCT_FMT)
        if ref:
            ws.cell(row, sc).fill = REF_FILL
            ws.cell(row, total_col).fill = REF_FILL
            if pct_col is not None:
                ws.cell(row, pct_col).fill = REF_FILL
        row += 2
    else:
        row += 1
    return row


def _build_hoja1_zonix_blocks(ws, start_row: int, dr: dict) -> int:
    """8 bloques Hoja1 tipo Pizza — datos Zonix vía espejo Detallado."""
    row = start_row
    pct = 9
    pc_start = dr["pc_start_row"]
    pc_end = dr["pc_end_row"]
    ade_end = dr["hq_adecuacion_end_row"]

    DETALLADO_ROWS["hoja1_equipos_row"] = row
    row = _mirror_detallado_block(
        ws,
        row,
        "equipos",
        pc_start,
        pc_end,
        subtotal_key=None,
        ref=True,
        pct_col=pct,
    )
    set_text(ws, row, 5, "Subtotal equipos (4 PCs)", True)
    set_num(ws, row, 8, formula=f"=SUM('{DET_SHEET}'!G{pc_start}:G{pc_end})")
    ws.cell(row, 5).font = SUB_FONT
    for col in (5, 8, 9):
        ws.cell(row, col).fill = REF_FILL
    DETALLADO_ROWS["hoja1_pc_subtotal_row"] = row
    row += 2

    DETALLADO_ROWS["hoja1_adecuacion_row"] = row
    row = _mirror_detallado_block(
        ws,
        row,
        "adecuacion",
        dr["hq_start_row"],
        ade_end,
        subtotal_key=None,
        ref=True,
        pct_col=pct,
    )
    set_text(ws, row, 5, "Subtotal adecuación HQ", True)
    set_num(
        ws,
        row,
        8,
        formula=f"=SUM('{DET_SHEET}'!G{dr['hq_start_row']}:G{ade_end})",
    )
    ws.cell(row, 5).font = SUB_FONT
    for col in (5, 8):
        ws.cell(row, col).fill = REF_FILL
    row += 2

    DETALLADO_ROWS["hoja1_mo_row"] = row
    row = _mirror_detallado_block(
        ws,
        row,
        "mano de obra",
        dr["mo_fase0_start_row"],
        dr["mo_fase0_end_row"],
        "subtotal_mo_fase0",
        ref=True,
        pct_col=pct,
    )

    DETALLADO_ROWS["hoja1_transporte_row"] = row
    row = _mirror_detallado_block(
        ws,
        row,
        "transporte",
        dr["transport_start_row"],
        dr["transport_end_row"],
        "subtotal_transporte_ref",
        ref=True,
        pct_col=pct,
    )

    DETALLADO_ROWS["hoja1_alquiler_row"] = row
    row = _mirror_detallado_block(
        ws,
        row,
        "alquiler",
        dr["hq_alquiler_row"],
        dr["hq_alquiler_row"],
        subtotal_key=None,
        ref=True,
        pct_col=pct,
    )
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=9)
    note = ws.cell(row, 5)
    note.value = (
        "Depósito HQ (one-shot) incluido en bloque adecuación. "
        "Arriendo mensual post-Day-D — ref. burn Detallado."
    )
    note.font = REF_FONT
    note.fill = REF_FILL
    note.alignment = Alignment(wrap_text=True)
    row += 2

    DETALLADO_ROWS["hoja1_legal_row"] = row
    row = _mirror_detallado_block(
        ws,
        row,
        "constitucion empresa y permisos",
        dr["legal_start_row"],
        dr["legal_end_row"],
        "subtotal_legal",
        ref=True,
        pct_col=pct,
    )

    DETALLADO_ROWS["hoja1_marketing_row"] = row
    row = _mirror_detallado_block(
        ws,
        row,
        "marketing — intro mercado / demo",
        dr["intro_start_row"],
        dr["intro_end_row"],
        "subtotal_intro",
        ref=True,
        pct_col=pct,
    )
    row = _mirror_detallado_block(
        ws,
        row,
        "marketing — pre-lanzamiento (3 meses)",
        dr["mkt_pre_start_row"],
        dr["mkt_pre_end_row"],
        "subtotal_mkt_pre",
        ref=True,
        pct_col=pct,
    )

    DETALLADO_ROWS["hoja1_materia_row"] = row
    hosting_r = dr["hosting_saas_row"]
    row = _mirror_detallado_block(
        ws,
        row,
        "materia (SaaS / hosting / SMS — marketplace)",
        hosting_r,
        hosting_r,
        subtotal_key=None,
        ref=True,
        pct_col=pct,
    )
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=9)
    foot = ws.cell(row, 5)
    foot.value = (
        "Equivalente digital marketplace (hosting/SMS). Desglose Detallado + "
        f"unit economics → ESTA SI VALE (~{ACTIVE_TIER.hosting} USD/mes)."
    )
    foot.font = REF_FONT
    foot.fill = REF_FILL
    foot.alignment = Alignment(wrap_text=True)
    row += 1
    return row


def _build_hoja1_nav(ws, start_row: int) -> int:
    """Enlaces a bloques Pizza restantes (Hoja3 / Detallado / ESTA)."""
    nav_items = [
        "Equipos producción: N/A — marketplace SaaS (Pizza equipos → no aplica)",
        "MO / legal / mkt / transporte: → Hoja3 Sección B",
        "Materia prima / SaaS mensual: → ESTA SI VALE (~154 USD/mes)",
        "Alquiler: depósito HQ (one-shot) + HQ ~500/mes en burn — ver Detallado",
    ]
    row = start_row
    for text in nav_items:
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=9)
        cell = ws.cell(row, 5)
        cell.value = text
        cell.font = REF_FONT
        cell.fill = REF_FILL
        cell.alignment = Alignment(wrap_text=True)
        row += 1
    return row


def _build_hoja1_panel_jk(ws, dr: dict) -> None:
    """Panel lateral J–K — refs Pizza inversión/mensual/reserva sin duplicar montos."""
    col_j, col_k = 10, 11

    set_text(ws, 1, col_j, f"RESUMEN CAPEX ({MODEL_VERSION})", True)
    ws.cell(1, col_j).font = TITLE_FONT

    ws.merge_cells(start_row=2, start_column=col_j, end_row=2, end_column=col_k)
    note = ws.cell(2, col_j)
    note.value = (
        "Refs informativas — HQ ya en one-shots; no sumar CapEx + TOTAL SAFE."
    )
    note.font = REF_FONT
    note.fill = REF_FILL
    note.alignment = Alignment(wrap_text=True)

    r = 3
    set_text(ws, r, col_j, "Inversión CapEx HQ")
    set_num(ws, r, col_k, formula=f"='{DET_SHEET}'!G{dr['subtotal_hq']}")
    DETALLADO_ROWS["hoja1_panel_inversion_row"] = r
    inv_row = r
    r += 1
    set_text(ws, r, col_j, "% CapEx / TOTAL SAFE Lean")
    set_num(
        ws,
        r,
        col_k,
        formula=f"=K{inv_row}/'{DET_SHEET}'!G{dr['total_safe']}",
        fmt=PCT_FMT,
    )
    r += 1
    set_text(ws, r, col_j, "TOTAL SAFE Lean (ref.)")
    set_num(ws, r, col_k, formula=f"='{DET_SHEET}'!G{dr['total_safe']}")
    DETALLADO_ROWS["hoja1_panel_total_safe_row"] = r
    total_safe_panel_row = r
    r += 1
    set_text(ws, r, col_j, "Burn mensual ref. (MO+mkt)")
    mo_r = dr["subtotal_mo_operativa"]
    mkt_r = dr["subtotal_mkt_mensual"]
    set_num(
        ws,
        r,
        col_k,
        formula=f"='{DET_SHEET}'!G{mo_r}+'{DET_SHEET}'!G{mkt_r}",
    )
    r += 1
    set_text(ws, r, col_j, "Reserva caja (ref.)")
    set_num(ws, r, col_k, formula=f"='{DET_SHEET}'!G{dr['reserva_row']}")
    r += 1
    set_text(ws, r, col_j, "% equity SAFE (ref.)")
    set_num(ws, r, col_k, formula=f"=K{total_safe_panel_row}/{CAP_LEAN}", fmt=PCT_FMT)
    r += 2
    set_text(ws, r, col_j, "Use of funds → Hoja3 Sección A")
    ws.cell(r, col_j).font = REF_FONT


def build_hoja3(ws):
    ws.title = "Hoja3"
    dr = DETALLADO_ROWS

    set_text(ws, 7, 5, f"Zonix Pharma — Resumen inversión ({MODEL_VERSION})", True)
    ws["E7"].font = TITLE_FONT

    # --- Sección A: use of funds (suma 100%) ---
    set_text(ws, 9, 5, "A — Use of funds (TOTAL SAFE Lean)", True)
    style_section_row(ws, 9, range(5, 9))
    set_text(ws, 10, 5, "Bloque", True)
    set_text(ws, 10, 7, "USD", True)
    set_text(ws, 10, 8, "% del total", True)
    style_header_row(ws, 10, [5, 7, 8])

    total_a_row = 15
    DETALLADO_ROWS["hoja3_total_a_row"] = total_a_row
    a_blocks = [
        ("One-shots (legal + intro + HQ)", dr["one_shots"]),
        ("Burn 12 meses (M1–M12)", dr["burn_row"]),
        ("Reserva caja Lean", dr["reserva_row"]),
    ]
    for i, (name, ref) in enumerate(a_blocks, start=11):
        set_text(ws, i, 5, name)
        set_num(ws, i, 7, formula=f"='{DET_SHEET}'!G{ref}")
        set_num(ws, i, 8, formula=f"=G{i}/G{total_a_row}", fmt=PCT_FMT)
    set_text(ws, total_a_row, 5, "TOTAL SAFE Lean", True)
    set_num(ws, total_a_row, 7, formula=f"='{DET_SHEET}'!G{dr['total_safe']}")
    set_num(ws, total_a_row, 8, formula=f"=SUM(H11:H13)", fmt=PCT_FMT)
    style_total_row(ws, total_a_row, cols=[5, 7, 8])
    set_text(
        ws,
        total_a_row + 1,
        5,
        "Marketing pre-lanzamiento (~2.250 USD) incluido en Fase 0 — ver Sección B. "
        "Unit economics → pestaña ESTA SI VALE.",
    )
    ws.cell(total_a_row + 1, 5).font = REF_FONT

    _build_resumen_lateral_km(ws, 9, include_timeline=True, link_detallado=True)

    # --- Sección B: desglose operativo tipo Pizza (enlace Detallado) ---
    row_b = total_a_row + 3
    set_text(ws, row_b, 5, "B — Desglose operativo Fase 0 (enlace Detallado)", True)
    style_section_row(ws, row_b, range(5, 9))
    row_b += 1
    ws.merge_cells(start_row=row_b, start_column=5, end_row=row_b, end_column=8)
    disclaimer = ws.cell(row_b, 5)
    disclaimer.value = (
        "Informativo — no sumar bloques al TOTAL SAFE (Sección A). "
        "Subtotal Fase 0 = modelo timing 0a–0c, no suma de columnas visibles."
    )
    disclaimer.font = REF_FONT
    disclaimer.fill = REF_FILL
    disclaimer.alignment = Alignment(wrap_text=True)
    row_b += 1
    set_text(ws, row_b, 5, "Descripción", True)
    set_text(ws, row_b, 6, "Cant.", True)
    set_text(ws, row_b, 7, "USD/u", True)
    set_text(ws, row_b, 8, "Total USD", True)
    style_header_row(ws, row_b, [5, 6, 7, 8])
    row_b += 1
    DETALLADO_ROWS["hoja3_section_b_start"] = row_b

    row_b = _mirror_block(
        ws,
        row_b,
        "Constitución empresa (sin BPF/CPE farmacia)",
        dr["legal_start_row"],
        dr["legal_end_row"],
        "subtotal_legal",
    )
    row_b = _mirror_block(
        ws,
        row_b,
        "Introducción al mercado / demo",
        dr["intro_start_row"],
        dr["intro_end_row"],
        "subtotal_intro",
    )
    row_b = _mirror_block(
        ws,
        row_b,
        "Marketing pre-lanzamiento (3 meses)",
        dr["mkt_pre_start_row"],
        dr["mkt_pre_end_row"],
        "subtotal_mkt_pre",
    )
    row_b = _mirror_block(
        ws,
        row_b,
        "Mano de obra Fase 0 (3 meses)",
        dr["mo_fase0_start_row"],
        dr["mo_fase0_end_row"],
        "subtotal_mo_fase0",
    )
    row_b = _mirror_block(
        ws,
        row_b,
        "Transporte B2B (referencia — no suma ancla)",
        dr["transport_start_row"],
        dr["transport_end_row"],
        "subtotal_transporte_ref",
        ref=True,
    )
    row_b = _mirror_block(
        ws,
        row_b,
        "HQ y CapEx",
        dr["hq_start_row"],
        dr["hq_end_row"],
        "subtotal_hq",
    )
    set_text(
        ws,
        row_b,
        5,
        "Materia prima: No aplica — marketplace SaaS (~154 USD/mes hosting/SMS en ESTA SI VALE).",
    )
    ws.cell(row_b, 5).font = REF_FONT
    ws.cell(row_b, 5).fill = REF_FILL
    row_b += 2
    set_text(ws, row_b, 5, "Validación: Fase 0 − one-shots − mkt pre (ops burn)", True)
    set_num(ws, row_b, 8, formula=f"='{DET_SHEET}'!G{dr['cross_ops_fase0']}")
    ws.cell(row_b, 5).font = SUB_FONT
    row_b += 2
    set_text(ws, row_b, 5, "Subtotal Fase 0 (T+0→Day-D)", True)
    set_num(ws, row_b, 8, formula=f"='{DET_SHEET}'!G{dr['subtotal_fase0']}")
    style_total_row(ws, row_b, cols=[5, 8])

    ws.column_dimensions["E"].width = 44
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["K"].width = 28
    ws.column_dimensions["M"].width = 14


def _esta_factor_row(ws, row: int, factor: float, note: str = "") -> tuple[int, str]:
    """Fila dedicada factor escalación Año 2 (patrón Pizza: J=constante antes del bloque)."""
    if note:
        ws.cell(row, 5).value = note
        ws.cell(row, 5).font = REF_FONT
    set_num(ws, row, 10, value=factor, fmt=PCT_FMT)
    ws.cell(row, 10).fill = REF_FILL
    ws.cell(row, 10).border = BORDER
    return row + 1, f"J${row}"


def _apply_esta_year2_col(ws, first_row: int, last_row: int, factor_cell: str) -> None:
    """Columna J ref. Año 2 — base col H (Total USD); factor en celda dedicada (sin circular)."""
    for r in range(first_row, last_row + 1):
        total = ws.cell(r, 8).value
        if isinstance(total, str) and DET_SHEET in total:
            set_num(ws, r, 10, formula=f"=H{r}*(1+${factor_cell})", fmt=NUM_FMT)
            ws.cell(r, 10).border = BORDER
            ws.cell(r, 10).fill = REF_FILL


def _apply_esta_year2_subtotal(
    ws, subtotal_row: int, first_row: int, last_row: int
) -> None:
    """Subtotal col J = SUM líneas detalle (paridad Pizza)."""
    set_num(
        ws,
        subtotal_row,
        10,
        formula=f"=SUM(J{first_row}:J{last_row})",
        fmt=NUM_FMT,
    )
    ws.cell(subtotal_row, 10).fill = REF_FILL
    ws.cell(subtotal_row, 10).border = BORDER


def _build_esta_investment_blocks(ws, start_row: int, dr: dict) -> int:
    """Bloques inversión ESTA (MO, transporte, legal, intro) — espejo Detallado."""
    row = start_row
    pct = 9

    row, mo_factor = _esta_factor_row(ws, row, 0.4, "Factor Año 2 MO (+40%)")
    DETALLADO_ROWS["esta_mo_row"] = row
    mo_title = row
    row = _mirror_detallado_block(
        ws,
        row,
        "Mano de obra (Fase 0 — 3 meses)",
        dr["mo_fase0_start_row"],
        dr["mo_fase0_end_row"],
        "subtotal_mo_fase0",
        ref=True,
        pct_col=pct,
    )
    mo_first, mo_last = mo_title + 1, row - 3
    _apply_esta_year2_col(ws, mo_first, mo_last, mo_factor)
    _apply_esta_year2_subtotal(ws, row - 2, mo_first, mo_last)

    row, tr_factor = _esta_factor_row(ws, row, 0.3, "Factor Año 2 transporte (+30%)")
    DETALLADO_ROWS["esta_transport_row"] = row
    tr_title = row
    row = _mirror_detallado_block(
        ws,
        row,
        "Transporte local (B2B / demo)",
        dr["transport_start_row"],
        dr["transport_end_row"],
        "subtotal_transporte_ref",
        ref=True,
        pct_col=pct,
    )
    tr_first, tr_last = tr_title + 1, row - 3
    _apply_esta_year2_col(ws, tr_first, tr_last, tr_factor)
    _apply_esta_year2_subtotal(ws, row - 2, tr_first, tr_last)

    row, legal_factor = _esta_factor_row(
        ws, row, 0.0, "Constitución one-shot (sin escalación Año 2)"
    )
    DETALLADO_ROWS["esta_legal_row"] = row
    legal_title = row
    row = _mirror_detallado_block(
        ws,
        row,
        "Constitución empresa y permisos (C.A. — sin BPF/CPE piloto)",
        dr["legal_start_row"],
        dr["legal_end_row"],
        "subtotal_legal",
        ref=True,
        pct_col=pct,
    )
    legal_first, legal_last = legal_title + 1, row - 3
    _apply_esta_year2_col(ws, legal_first, legal_last, legal_factor)
    _apply_esta_year2_subtotal(ws, row - 2, legal_first, legal_last)

    row, intro_factor = _esta_factor_row(
        ws, row, -0.5, "Factor Año 2 intro/marketing (-50% one-shot)"
    )
    DETALLADO_ROWS["esta_intro_row"] = row
    intro_title = row
    row = _mirror_detallado_block(
        ws,
        row,
        "Introducción al mercado; primeros 3 meses",
        dr["intro_start_row"],
        dr["intro_end_row"],
        "subtotal_intro",
        ref=True,
        pct_col=pct,
    )
    intro_first = intro_title + 1
    row = _mirror_detallado_block(
        ws,
        row,
        "Marketing pre-lanzamiento (3 meses)",
        dr["mkt_pre_start_row"],
        dr["mkt_pre_end_row"],
        "subtotal_mkt_pre",
        ref=True,
        pct_col=pct,
    )
    intro_last = row - 3
    _apply_esta_year2_col(ws, intro_first, intro_last, intro_factor)
    _apply_esta_year2_subtotal(ws, row - 2, intro_first, intro_last)

    intro_sub = dr.get("subtotal_intro")
    mkt_sub = dr.get("subtotal_mkt_pre")
    if intro_sub and mkt_sub:
        set_text(ws, row, 5, "Promedio mensual intro + mkt pre (3 meses)", True)
        set_num(
            ws,
            row,
            8,
            formula=f"=('{DET_SHEET}'!G{intro_sub}+'{DET_SHEET}'!G{mkt_sub})/3",
        )
        set_num(ws, row, 10, formula=f"=H{row}*(1+${intro_factor})")
        ws.cell(row, 5).font = SUB_FONT
        for col in (5, 8, 10):
            ws.cell(row, col).fill = REF_FILL
        DETALLADO_ROWS["esta_intro_mensual_row"] = row
        row += 1

    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=10)
    note = ws.cell(row, 5)
    note.value = (
        "CapEx HQ / equipos / SaaS mensual → Hoja1 + Detallado. "
        "Use-of-funds SAFE → Hoja3 §A. No sumar bloques ESTA al TOTAL SAFE."
    )
    note.font = REF_FONT
    note.fill = REF_FILL
    note.alignment = Alignment(wrap_text=True)
    row += 1
    return row


def _build_esta_gastos_fijos_saas(ws, dr: dict) -> None:
    """Equivalente Pizza «gastos fijos mensuales» — bandas margen plataforma SaaS (col P, junto panel M)."""
    intro_m = dr.get("esta_intro_mensual_row")
    burn_total = dr.get("esta_burn_total_row")
    if not burn_total:
        return
    act_row = YEAR_ROWS.get(1, {}).get("activas", 18)

    set_text(ws, 8, 16, "gastos fijos mensuales (ref. burn/activas — no sumar SAFE)", True)
    ws.cell(8, 16).font = Font(bold=True, size=9)
    ws.cell(8, 16).alignment = Alignment(wrap_text=True)

    scenarios = [
        (12, "Escenario margen 50%-30%"),
        (17, "Escenario margen 60%-40%"),
        (22, "Escenario margen 70%-50%"),
    ]
    for anchor_row, label in scenarios:
        set_text(ws, anchor_row, 16, label, True)
        set_text(ws, anchor_row + 1, 16, "Mes 1")
        set_text(ws, anchor_row + 2, 16, "Mes 2")
        set_text(ws, anchor_row + 3, 16, "Mes 3")
        set_text(ws, anchor_row + 4, 16, "TOTAL ref.")
        burn_ref = f"='ESTA SI VALE'!$I${burn_total}"
        act_ref = f"'Año 1'!D{act_row}"
        m1_formula = f"=IF({act_ref}=0,0,{burn_ref}/{act_ref})"
        for i, off in enumerate((1, 2, 3)):
            r = anchor_row + off
            if i == 0:
                set_num(ws, r, 17, formula=m1_formula)
            elif intro_m:
                set_num(ws, r, 17, formula=f"=H{intro_m}")
            else:
                set_num(ws, r, 17, formula=burn_ref)
        set_num(
            ws,
            anchor_row + 4,
            17,
            formula=f"=SUM(Q{anchor_row + 1}:Q{anchor_row + 3})",
        )
        for r in range(anchor_row, anchor_row + 5):
            ws.cell(r, 16).fill = REF_FILL
            ws.cell(r, 17).fill = REF_FILL


def _build_esta_unit_panel(ws, start_col: int = 13) -> int:
    """Unit economics + tiers + P10/P50 — panel lateral (cols M+)."""
    c_concept = start_col
    c_val = start_col + 2
    c_extra1 = start_col + 3
    c_extra2 = start_col + 4
    c_src = start_col + 6

    set_text(ws, 7, c_concept, "Zonix Pharma — Unit economics y escenarios", True)
    ws.cell(7, c_concept).font = TITLE_FONT

    unit_rows = [
        ("ARPF placeholder (USD/mes)", 50, "UNIT_ECONOMICS"),
        ("CAC farmacia (USD)", 139, "UNIT_ECONOMICS"),
        ("Churn mensual target", 0.05, "UNIT_ECONOMICS"),
        ("LTV (USD)", 1000, "UNIT_ECONOMICS"),
        ("LTV / CAC", 7.2, "UNIT_ECONOMICS"),
        ("Payback CAC (meses)", 2.8, "UNIT_ECONOMICS"),
        ("Margen bruto plataforma", 0.92, "UNIT_ECONOMICS"),
        ("Cuota Basic (USD/mes)", 25, "PROPUESTA B2B §5"),
        ("Cuota Pro (USD/mes)", 40, "PROPUESTA B2B §5"),
        ("Cuota Enterprise (USD/mes)", 55, "PROPUESTA B2B §5"),
        ("Mix Basic / Pro / Enterprise", "60/30/10", "PROYECCION §1.2"),
        ("Farmacias activas equilibrio Lean (M12 v3.3)", EQUILIBRIO_ACTIVAS_LEAN, "PROYECCION — 4× Sales"),
    ]
    set_text(ws, 9, c_concept, "Concepto", True)
    set_text(ws, 9, c_val, "Valor", True)
    set_text(ws, 9, c_src, "Fuente", True)
    style_header_row(ws, 9, [c_concept, c_val, c_src])
    r = 10
    for label, val, src in unit_rows:
        set_text(ws, r, c_concept, label)
        if isinstance(val, float) and val < 1:
            set_num(ws, r, c_val, value=val, fmt=PCT_FMT)
        elif isinstance(val, (int, float)):
            set_num(ws, r, c_val, value=val)
        else:
            set_text(ws, r, c_val, str(val))
        set_text(ws, r, c_src, src)
        r += 1

    r += 1
    set_text(ws, r, c_concept, "ESCENARIOS CAPITAL (tiers)", True)
    style_section_row(ws, r, range(c_concept, c_extra2 + 1))
    r += 1
    set_text(ws, r, c_concept, "Tier")
    set_text(ws, r, c_val, "Capital")
    set_text(ws, r, c_extra1, "Burn prom.")
    set_text(ws, r, c_extra2, "Activas eq.")
    set_text(ws, r, c_src, "SAFE cap")
    r += 1
    for key in ("lean", "lean_plus", "base", "growth", "blitz"):
        s = TIER_SUMMARY[key]
        set_text(ws, r, c_concept, f"{TIERS[key].label} — capital (USD)")
        set_num(ws, r, c_val, value=s["capital"])
        set_num(ws, r, c_extra1, value=s["burn_avg"])
        set_num(ws, r, c_extra2, value=s["activas_eq"])
        set_num(ws, r, c_src, value=s["safe_cap"])
        r += 1

    r += 1
    set_text(ws, r, c_concept, "ESCENARIOS PROYECCION §7 (P10/P50/P90)", True)
    style_section_row(ws, r, range(c_concept, c_extra2 + 1))
    r += 1
    set_text(ws, r, c_concept, "Escenario")
    set_text(ws, r, c_val, "Activas M12")
    set_text(ws, r, c_extra1, "ARPF")
    set_text(ws, r, c_extra2, "Cash M12")
    r += 1
    p50_cash = round(CAJA_M12_THEORETICAL)
    lean_plus = TIER_SUMMARY["lean_plus"]
    for esc, act, arpf, cash in [
        ("P10 pesimista", P10_ACTIVAS_M12, 40, P10_CASH_M12),
        ("P50 base", ACTIVAS[-1], 50, p50_cash),
        ("P90 optimista", P90_ACTIVAS_M12, 55, P90_CASH_M12),
        ("Lean+ stretch (4×, curva ×1,15)", lean_plus["activas_eq"], 50, max(0, lean_plus["caja_m12"])),
        (
            "Blitz stretch (6× Sales)",
            TIER_SUMMARY["blitz"]["activas_eq"],
            50,
            max(0, TIER_SUMMARY["blitz"]["caja_m12"]),
        ),
    ]:
        set_text(ws, r, c_concept, esc)
        set_num(ws, r, c_val, value=act)
        set_num(ws, r, c_extra1, value=arpf)
        if isinstance(cash, (int, float)):
            set_num(ws, r, c_extra2, value=cash)
        else:
            set_text(ws, r, c_extra2, str(cash))
        r += 1
    return r


def _build_esta_burn_catalog(ws, start_row: int) -> int:
    """Catálogo burn USD/mes — ancla fija col I para enlaces Año 1."""
    r = start_row
    set_text(ws, r, 5, "BURN MENSUAL REF. (USD/mes) — enlazado Año 1", True)
    style_section_row(ws, r, range(5, 12))
    r += 1
    set_text(ws, r, 5, "Concepto")
    set_text(ws, r, 9, "USD/mes")
    set_text(ws, r, 11, "Fuente")
    r += 1
    t = ACTIVE_TIER
    meta_avg = round(sum(META_BY_MONTH) / 12)
    valla_avg = round(sum(VALLA_BY_MONTH) / 12)
    cont_avg = round(sum(CONTINGENCIA_BY_MONTH) / 12)
    burn_catalog = burn_catalog_for_tier(t, meta_avg, valla_avg, cont_avg)
    burn_start = r
    DETALLADO_ROWS["esta_burn_start"] = burn_start
    for label, val, src in burn_catalog:
        set_text(ws, r, 5, label)
        set_num(ws, r, 9, value=val)
        if src:
            set_text(ws, r, 11, src)
        ESTA_BURN_ROWS[label] = r
        if label == t.dev_label:
            ESTA_BURN_ROWS[t.dev_label] = r
        sales_key = f"{t.sales_count}× Sales B2B"
        if sales_key in label:
            ESTA_BURN_ROWS[sales_key] = r
        if label == "Servicios oficina":
            ESTA_BURN_ROWS["Servicios oficina (electricidad, agua, internet)"] = r
        if "Valla mediana" in label:
            ESTA_BURN_ROWS["Valla mediana (desde M2)"] = r
        if "Valla pequeña" in label:
            ESTA_BURN_ROWS["Valla pequeña (desde M3)"] = r
        if "contingencia" in label:
            ESTA_BURN_ROWS["Material + transporte + contingencia (incl. buffer VE 20%)"] = r
        r += 1
    set_text(ws, r, 5, f"Sales comisiones (estim. prom/mes — ${t.sales_commission_per_signing}/firma)")
    set_num(ws, r, 9, value=SALES_COMMISSION_M_AVG)
    set_text(ws, r, 11, "Variable — no en burn fijo Lean")
    ws.cell(r, 5).fill = REF_FILL
    ws.cell(r, 5).font = REF_FONT
    r += 1
    contingencia_row = burn_start + len(burn_catalog) - 1
    fixed_end_row = contingencia_row - 1
    set_text(ws, r, 5, "Total burn ref. catálogo (fijos + contingencia)", True)
    set_num(
        ws,
        r,
        9,
        formula=f"=SUM(I{burn_start}:I{fixed_end_row})+I{contingencia_row}",
    )
    DETALLADO_ROWS["esta_burn_total_row"] = r
    r += 1
    set_text(ws, r, 5, f"Payback inversor ilustrativo (SAFE {round(EQUITY_INV * 100, 2)}%)", True)
    set_text(ws, r + 1, 5, "Ver hoja Flujo Total — fila Payback (años). No promesa contractual.")
    r += 2
    set_text(ws, r, 5, f"Modelo financiero {MODEL_VERSION} — jun 2026 — pack Lanzamiento Zonix Pharma")
    ws.cell(r, 5).font = REF_FONT
    return r


def populate_esta_simulator(ws, wb: Workbook) -> None:
    """Grid M1-M12 marketplace + gastos fijos SaaS — enlaces Año 1."""
    if 1 not in YEAR_ROWS:
        return
    _build_esta_gastos_fijos_saas(ws, DETALLADO_ROWS)
    yr = YEAR_ROWS[1]
    act_row = yr["activas"]
    ing_row = yr["ing"]
    cost_row = yr["cost"]
    fcf_row = yr["fcf"]
    start = DETALLADO_ROWS.get("esta_sim_anchor_row")
    if not start:
        return
    DETALLADO_ROWS["esta_sim_start"] = start

    set_text(ws, start, 15, "SIMULADOR MARKETPLACE (espejo Año 1 — P50 base)", True)
    ws.cell(start, 15).font = Font(bold=True, size=10)
    hdr = start + 1
    headers = [
        (15, "Mes"),
        (16, "Costo/activa (burn÷activas)"),
        (17, "Farmacias activas"),
        (18, "Costo total mes"),
        (19, "ARPF (USD)"),
        (20, "Revenue mes"),
        (21, "Util./activa"),
        (22, "Utilidad mes"),
        (23, "Escenario"),
    ]
    for col, label in headers:
        set_text(ws, hdr, col, label, True)
        style_header_row(ws, hdr, [col])

    data_start = hdr + 1
    for m in range(12):
        r = data_start + m
        mc = col_month(m + 1)
        cl = get_column_letter(mc)
        set_num(ws, r, 15, value=m + 1)
        set_num(ws, r, 17, formula=f"='Año 1'!{cl}{act_row}")
        set_num(ws, r, 16, formula=f"=IF(Q{r}=0,0,'Año 1'!{cl}{cost_row}/Q{r})")
        set_num(ws, r, 18, formula=f"='Año 1'!{cl}{cost_row}")
        set_num(ws, r, 19, formula=f"=IF(Q{r}=0,0,'Año 1'!{cl}{ing_row}/Q{r})")
        set_num(ws, r, 20, formula=f"='Año 1'!{cl}{ing_row}")
        set_num(ws, r, 21, formula=f"=S{r}-P{r}")
        set_num(ws, r, 22, formula=f"='Año 1'!{cl}{fcf_row}")
        set_text(ws, r, 23, "P50 base")
        for col in range(15, 24):
            ws.cell(r, col).fill = REF_FILL

    sum_row = data_start + 12
    set_text(ws, sum_row, 15, "Total utilidad M1-M12 (P50)", True)
    set_num(
        ws,
        sum_row,
        22,
        formula=f"=SUM(V{data_start}:V{data_start + 11})",
    )

    p10_start = sum_row + 2
    set_text(ws, p10_start, 15, "Stress P10 (ARPF 40 fijo)", True)
    hdr2 = p10_start + 1
    set_text(ws, hdr2, 15, "Mes")
    set_text(ws, hdr2, 17, "Activas")
    set_text(ws, hdr2, 19, "ARPF")
    set_text(ws, hdr2, 20, "Revenue")
    set_text(ws, hdr2, 18, "Costo")
    set_text(ws, hdr2, 22, "Utilidad")
    for m in range(12):
        r = hdr2 + 1 + m
        mc = col_month(m + 1)
        cl = get_column_letter(mc)
        set_num(ws, r, 15, value=m + 1)
        set_num(ws, r, 17, formula=f"='Año 1'!{cl}{act_row}")
        set_num(ws, r, 19, value=40)
        set_num(ws, r, 20, formula=f"=Q{r}*S{r}")
        set_num(ws, r, 18, formula=f"='Año 1'!{cl}{cost_row}")
        set_num(ws, r, 22, formula=f"=T{r}-R{r}")
        ws.cell(r, 23).value = "P10 stress"

    set_text(ws, hdr2 + 13, 15, "Total utilidad P10", True)
    set_num(
        ws,
        hdr2 + 13,
        22,
        formula=f"=SUM(V{hdr2 + 1}:V{hdr2 + 12})",
    )

    for col_letter, width in (
        ("E", 44),
        ("F", 8),
        ("G", 10),
        ("H", 14),
        ("I", 10),
        ("J", 12),
        ("M", 28),
        ("O", 12),
        ("P", 14),
        ("Q", 12),
        ("R", 12),
        ("S", 10),
        ("T", 12),
        ("U", 10),
        ("V", 12),
        ("W", 12),
    ):
        ws.column_dimensions[col_letter].width = width


def build_esta_si_vale(ws):
    ws.title = "ESTA SI VALE"
    dr = DETALLADO_ROWS

    ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=10)
    disclaimer = ws.cell(3, 5)
    disclaimer.value = (
        "Datos ZonixPharma — espejo Detallado/Hoja1. Bloques ESTA = vista due diligence. "
        "TOTAL SAFE oficial → Hoja3 §A. No sumar bloques al SAFE."
    )
    disclaimer.font = REF_FONT
    disclaimer.fill = REF_FILL
    disclaimer.alignment = Alignment(wrap_text=True)

    set_text(ws, 7, 5, f"Zonix Pharma — ESTA SI VALE (espejo Pizza {MODEL_VERSION})", True)
    ws["E7"].font = TITLE_FONT

    set_text(ws, 9, 5, "Descripción", True)
    set_text(ws, 9, 6, "Cant.", True)
    set_text(ws, 9, 7, "USD/u", True)
    set_text(ws, 9, 8, "Total USD", True)
    set_text(ws, 9, 9, "% SAFE", True)
    set_text(ws, 9, 10, "Año 2 ref.", True)
    style_header_row(ws, 9, [5, 6, 7, 8, 9, 10])

    next_row = _build_esta_investment_blocks(ws, 10, dr)
    _build_esta_unit_panel(ws, start_col=13)
    burn_start = next_row + ESTA_ROW_GAP
    DETALLADO_ROWS["esta_burn_anchor_row"] = burn_start
    DETALLADO_ROWS["esta_sim_anchor_row"] = burn_start
    _build_esta_burn_catalog(ws, burn_start)


def build_year_sheet(ws, year_num: int, wb: Workbook):
    ws.title = f"Año {year_num}"
    yd = YEAR_DATA[year_num - 1]
    _, activas_end, rev_y, cost_y, note = yd

    # Fila 2: encabezado único ITEMS / DATOS (layout Pizza)
    set_text(ws, 2, 2, "ITEMS", True)
    set_text(ws, 2, 3, "DATOS", True)
    style_header_row(ws, 2, [2, 3])

    equity_row = None
    prem_rows: dict[str, int] = {}

    if year_num == 1:
        premises = [
            ("ARPF placeholder USD/mes", 50, "arpf"),
            ("Cuota Basic USD/mes", 25, "basic"),
            ("Cuota Pro USD/mes", 40, "pro"),
            ("Cuota Enterprise USD/mes", 55, "ent"),
            ("Mix Basic (ponderador)", 0.6, "mix_b"),
            ("Mix Pro (ponderador)", 0.3, "mix_p"),
            ("Mix Enterprise (ponderador)", 0.1, "mix_e"),
            ("Churn mensual target", 0.05, None),
            ("Participación Inversor SAFE", EQUITY_INV, "equity"),
            ("Participación Founders", None, None),
            ("Tasa descuento VAN", DISCOUNT, None),
            ("Capital SAFE Lean (USD)", SAFE_LEAN, None),
            ("Nota escenario", "Lean ancla pack", None),
        ]
        pr = 3
        for label, val, key in premises:
            set_text(ws, pr, 2, label)
            if label == "Participación Inversor SAFE":
                equity_row = pr
                set_num(ws, pr, 3, value=val, fmt=PCT_FMT)
            elif label == "Participación Founders":
                set_num(ws, pr, 3, formula=f"=1-C{equity_row}", fmt=PCT_FMT)
            elif isinstance(val, float) and val < 1 and "Mix" not in label:
                set_num(ws, pr, 3, value=val, fmt=PCT_FMT)
            elif isinstance(val, (int, float)):
                set_num(ws, pr, 3, value=val)
            else:
                set_text(ws, pr, 3, str(val))
            if key:
                prem_rows[key] = pr
            pr += 1
    else:
        prev = f"Año {year_num - 1}"
        prev_act = YEAR_ROWS.get(year_num - 1, {}).get("activas", 18)
        premises = [
            ("Farmacias activas cierre (ref.)", f"='{prev}'!N{prev_act}", None),
            ("Revenue anual USD [LARGO PLAZO]", rev_y, None),
            ("Costos anual USD [LARGO PLAZO]", cost_y, None),
            ("Participación Inversor SAFE", EQUITY_INV, "equity"),
            ("Participación Founders", None, None),
            ("Etiqueta", note, None),
        ]
        pr = 3
        for label, val, key in premises:
            set_text(ws, pr, 2, label)
            if label == "Participación Inversor SAFE":
                equity_row = pr
                set_num(ws, pr, 3, value=val, fmt=PCT_FMT)
            elif label == "Participación Founders":
                set_num(ws, pr, 3, formula=f"=1-C{equity_row}", fmt=PCT_FMT)
            elif isinstance(val, str) and val.startswith("="):
                set_num(ws, pr, 3, formula=val)
            elif isinstance(val, (int, float)):
                set_num(ws, pr, 3, value=val)
            else:
                set_text(ws, pr, 3, str(val))
            if key:
                prem_rows[key] = pr
            pr += 1

    year_label_row = pr
    set_text(ws, year_label_row, 2, f"AÑO {year_num}", True)
    ws.cell(year_label_row, 2).font = TITLE_FONT
    header_row = year_label_row + 1
    month_headers(ws, header_row)

    r = header_row + 1
    if year_num == 1:
        set_text(ws, r, 2, "Farmacias firmadas (altas mes)")
        fill_month_row(ws, r, FIRMADAS)
        firmadas_row = r
        r += 1
        set_text(ws, r, 2, "Farmacias activas (stock cierre mes)")
        fill_month_row(ws, r, ACTIVAS, total_formula=f"=N{r}")
        activas_row = r
        r += 2
    else:
        set_text(ws, r, 2, "Farmacias activas promedio mes [LARGO PLAZO]")
        avg = int(activas_end * 0.85)
        fill_month_row(ws, r, [avg] * 12, total_formula=f"=N{r}")
        activas_row = r
        firmadas_row = r
        r += 2

    set_text(ws, r, 2, "Ingresos ($)", True)
    style_section_row(ws, r)
    r += 1

    if year_num == 1:
        ar = activas_row
        cuota_row = r
        set_text(ws, r, 2, "      Ingresos cuota fija B2B")
        for m in range(12):
            c = col_month(m + 1)
            cl = get_column_letter(c)
            set_num(
                ws, r, c,
                formula=(
                    f"={cl}{ar}*($C${prem_rows['mix_b']}*$C${prem_rows['basic']}"
                    f"+$C${prem_rows['mix_p']}*$C${prem_rows['pro']}"
                    f"+$C${prem_rows['mix_e']}*$C${prem_rows['ent']})"
                ),
            )
        sc, ec = get_column_letter(col_month(1)), get_column_letter(col_month(12))
        set_num(ws, r, col_total(), formula=f"=SUM({sc}{r}:{ec}{r})")
        r += 1

        gmv_row = r
        set_text(ws, r, 2, "      Ingresos comisión % GMV [EDITAR]")
        for m in range(12):
            c = col_month(m + 1)
            cl = get_column_letter(c)
            set_num(ws, r, c, formula=f"={cl}{ar}*($C${prem_rows['arpf']}-{cl}{cuota_row}/{cl}{ar})")
        set_num(ws, r, col_total(), formula=f"=SUM({sc}{gmv_row}:{ec}{gmv_row})")
        r += 1

        adj_row = r
        set_text(ws, r, 2, "      Ajuste billing pack PROYECCION §1.1")
        for m in range(12):
            c = col_month(m + 1)
            cl = get_column_letter(c)
            set_num(ws, r, c, formula=f"={REVENUE[m]}-({cl}{cuota_row}+{cl}{gmv_row})")
        set_num(ws, r, col_total(), formula=f"=SUM({sc}{r}:{ec}{r})")
        r += 1

        set_text(ws, r, 2, "Total Ingresos", True)
        ing_row = r
        for m in range(12):
            c = col_month(m + 1)
            cl = get_column_letter(c)
            set_num(ws, r, c, formula=f"={cl}{cuota_row}+{cl}{gmv_row}+{cl}{adj_row}")
        set_num(ws, r, col_total(), formula=f"=SUM({sc}{r}:{ec}{r})")
        style_total_row(ws, r)
        r += 2
    else:
        rev_vals = quarterly_monthly(year_num, "rev")
        set_text(ws, r, 2, "      Revenue plataforma")
        fill_month_row(ws, r, rev_vals)
        rev_detail_row = r
        r += 1
        set_text(ws, r, 2, "Total Ingresos", True)
        ing_row = r
        sc, ec = get_column_letter(col_month(1)), get_column_letter(col_month(12))
        for m in range(12):
            c = col_month(m + 1)
            cl = get_column_letter(c)
            set_num(ws, r, c, formula=f"={cl}{rev_detail_row}")
        set_num(ws, r, col_total(), formula=f"=SUM({sc}{r}:{ec}{r})")
        style_total_row(ws, r)
        r += 2

    set_text(ws, r, 2, "Costos ($)", True)
    style_section_row(ws, r)
    r += 1
    cost_start = r

    if year_num == 1:
        expense_row_map = {}
        for line_name, esta_key in EXPENSE_LINES:
            set_text(ws, r, 2, f"      {line_name}")
            if esta_key and esta_key in ESTA_BURN_ROWS:
                er = ESTA_BURN_ROWS[esta_key]
                for m in range(12):
                    c = col_month(m + 1)
                    set_num(ws, r, c, formula=f"='ESTA SI VALE'!$I${er}")
            else:
                vals = [expense_value(line_name, i) for i in range(12)]
                fill_month_row(ws, r, vals)
            expense_row_map[line_name] = r
            r += 1

        set_text(ws, r, 2, "Total Costos", True)
        cost_end = r
        sc = get_column_letter(col_month(1))
        ec = get_column_letter(col_month(12))
        for m in range(12):
            c = col_month(m + 1)
            parts = "+".join(f"{get_column_letter(c)}{expense_row_map[n]}" for n, _ in EXPENSE_LINES)
            set_num(ws, r, c, formula=f"={parts}")
        set_num(ws, r, col_total(), formula=f"=SUM({sc}{r}:{ec}{r})")
        style_total_row(ws, r)
        r += 2
    else:
        cost_vals = quarterly_monthly(year_num, "cost")
        cost_lines = [
            ("Equipo comercial + ops", [round(v * 0.35, 2) for v in cost_vals]),
            ("Marketing digital", [round(v * 0.18, 2) for v in cost_vals]),
            ("Hosting + SaaS", [round(v * 0.08, 2) for v in cost_vals]),
            ("Legal / contador", [round(v * 0.07, 2) for v in cost_vals]),
            ("HQ + servicios", [round(v * 0.12, 2) for v in cost_vals]),
            ("Imprevistos / buffer", [round(v * 0.20, 2) for v in cost_vals]),
        ]
        for label, vals in cost_lines:
            set_text(ws, r, 2, f"      {label}")
            fill_month_row(ws, r, vals)
            r += 1
        set_text(ws, r, 2, "Total Costos", True)
        cost_end = r
        sc, ec = get_column_letter(col_month(1)), get_column_letter(col_month(12))
        fill_month_row(ws, r, None, total_formula=f"=SUM({sc}{cost_start}:{ec}{r-1})")
        style_total_row(ws, r)
        r += 2

    set_text(ws, r, 2, "FCF", True)
    style_section_row(ws, r)
    fcf_row = r
    for m in range(12):
        c = col_month(m + 1)
        cl = get_column_letter(c)
        set_num(ws, r, c, formula=f"={cl}{ing_row}-{cl}{cost_end}")
    sc, ec = get_column_letter(col_month(1)), get_column_letter(col_month(12))
    set_num(ws, r, col_total(), formula=f"=SUM({sc}{r}:{ec}{r})")
    style_total_row(ws, r)
    r += 1

    set_text(ws, r, 2, "FCF acumulado")
    fcf_acum_row = r
    for m in range(12):
        c = col_month(m + 1)
        cl = get_column_letter(c)
        if m == 0:
            set_num(ws, r, c, formula=f"={cl}{fcf_row}")
        else:
            prev = get_column_letter(col_month(m))
            set_num(ws, r, c, formula=f"={prev}{r}+{cl}{fcf_row}")
    set_num(ws, r, col_total(), formula=f"={get_column_letter(col_month(12))}{r}")
    r += 1

    set_text(ws, r, 2, f"CCF (inversor SAFE {_equity_pct_es()}%)")
    for m in range(12):
        c = col_month(m + 1)
        cl = get_column_letter(c)
        set_num(ws, r, c, formula=f"={cl}{fcf_row}*$C${equity_row}")
    r += 1

    set_text(ws, r, 2, f"ECF (founders {_founders_pct_es()}%)")
    for m in range(12):
        c = col_month(m + 1)
        cl = get_column_letter(c)
        set_num(ws, r, c, formula=f"={cl}{fcf_row}-{cl}{r-1}")

    cash_ini_row = None
    cash_final_row = None
    if year_num == 1:
        r += 2
        set_text(ws, r, 2, "Cash deployment (post Day-D)", True)
        style_section_row(ws, r)
        r += 1
        cash_ini_row = r
        set_text(ws, r, 2, "Cash inicial mes")
        caja_det = DETALLADO_ROWS["caja_dayd"]
        set_num(
            ws, r, col_month(1),
            formula=f"='Detallado de la inversión.'!G{caja_det}",
        )
        r += 1
        cash_final_row = r
        set_text(ws, r, 2, "Cash final mes")
        for m in range(12):
            c = col_month(m + 1)
            cl = get_column_letter(c)
            set_num(ws, r, c, formula=f"={cl}{cash_ini_row}+{cl}{ing_row}-{cl}{cost_end}")
        for m in range(2, 13):
            c = col_month(m)
            prev = get_column_letter(col_month(m - 1))
            set_num(ws, cash_ini_row, c, formula=f"={prev}{cash_final_row}")

    YEAR_ROWS[year_num] = {
        "activas": activas_row,
        "firmadas": firmadas_row,
        "ing": ing_row,
        "cost": cost_end,
        "fcf": fcf_row,
        "fcf_acum": fcf_acum_row,
        "equity": equity_row,
        "cash_ini": cash_ini_row,
        "cash_final": cash_final_row,
    }

    ws.column_dimensions["B"].width = 42
    for c in range(3, 16):
        ws.column_dimensions[get_column_letter(c)].width = 11


def build_flujo_total(wb: Workbook):
    ws = wb.create_sheet("Flujo Total")
    set_text(ws, 2, 2, "PREMISAS DEL PROYECTO E INDICADORES — Zonix Pharma", True)
    ws["B2"].font = TITLE_FONT
    set_text(ws, 4, 2, "ITEMS", True)
    set_text(ws, 4, 3, "DATOS", True)
    style_header_row(ws, 4, [2, 3])

    prem = [
        ("ARPF placeholder USD/mes", 50),
        ("Participación Inversor SAFE", EQUITY_INV),
        ("Participación Founders", "=1-C6"),
        ("Tasa Descuento", DISCOUNT),
        ("Inversión Inicial (wire T+0)", -SAFE_LEAN),
        ("SAFE post-money cap Lean", CAP_LEAN),
        ("Nota reparto", "Ilustrativo — no sustituye term sheet SAFE"),
    ]
    for i, (label, val) in enumerate(prem, start=5):
        set_text(ws, i, 2, label)
        if isinstance(val, str) and val.startswith("="):
            set_num(ws, i, 3, formula=val, fmt=PCT_FMT)
        elif isinstance(val, float) and val < 0:
            set_num(ws, i, 3, value=val)
        elif isinstance(val, float) and val < 1:
            set_num(ws, i, 3, value=val, fmt=PCT_FMT)
        elif isinstance(val, (int, float)):
            set_num(ws, i, 3, value=val)
        else:
            set_text(ws, i, 3, val)

    set_text(ws, 12, 11, "CCF")
    set_text(ws, 12, 12, "Por recuperar")

    rr = 16
    for y in range(1, 6):
        set_text(ws, rr, y + 3, f"Año {y}")
    set_text(ws, rr, 9, "TOTAL")
    style_header_row(ws, rr, range(3, 10))

    rr += 1
    set_text(ws, rr, 2, "Farmacias activas cierre")
    for y in range(1, 6):
        ref = YEAR_ROWS[y]["activas"]
        set_num(ws, rr, y + 3, formula=f"='Año {y}'!N{ref}")
    set_num(ws, rr, 9, formula=f"=H{rr}")
    activas_flujo_row = rr
    rr += 1

    set_text(ws, rr, 2, "Total Ingresos anual")
    for y in range(1, 6):
        ref = YEAR_ROWS[y]["ing"]
        set_num(ws, rr, y + 3, formula=f"='Año {y}'!O{ref}")
    set_num(ws, rr, 9, formula=f"=SUM(D{rr}:H{rr})")
    ing_flujo_row = rr
    rr += 1

    set_text(ws, rr, 2, "Total Costos anual")
    for y in range(1, 6):
        ref = YEAR_ROWS[y]["cost"]
        set_num(ws, rr, y + 3, formula=f"='Año {y}'!O{ref}")
    set_num(ws, rr, 9, formula=f"=SUM(D{rr}:H{rr})")
    cost_flujo_row = rr
    rr += 1

    set_text(ws, rr, 2, "FCF anual", True)
    style_section_row(ws, rr)
    fcf_flujo_row = rr
    for y in range(1, 6):
        c = y + 3
        cl = get_column_letter(c)
        set_num(ws, rr, c, formula=f"={cl}{ing_flujo_row}-{cl}{cost_flujo_row}")
    set_num(ws, rr, 9, formula=f"=SUM(D{rr}:H{rr})")
    style_total_row(ws, rr)
    rr += 1

    set_text(ws, rr, 2, "FCF acumulado")
    for y in range(1, 6):
        c = y + 3
        cl = get_column_letter(c)
        if y == 1:
            set_num(ws, rr, c, formula=f"={cl}{fcf_flujo_row}")
        else:
            prev = get_column_letter(c - 1)
            set_num(ws, rr, c, formula=f"={prev}{rr}+{cl}{fcf_flujo_row}")
    fcf_acum_flujo_row = rr
    rr += 1

    set_text(ws, rr, 2, "CCF (inversor SAFE)")
    ccf_flujo_row = rr
    for y in range(1, 6):
        c = y + 3
        cl = get_column_letter(c)
        set_num(ws, rr, c, formula=f"={cl}{fcf_flujo_row}*$C$6")
    rr += 1

    set_text(ws, rr, 2, "CCF acumulado (inversor)")
    ccf_acum_flujo_row = rr
    for y in range(1, 6):
        c = y + 3
        cl = get_column_letter(c)
        if y == 1:
            set_num(ws, rr, c, formula=f"={cl}{ccf_flujo_row}")
        else:
            prev = get_column_letter(c - 1)
            set_num(ws, rr, c, formula=f"={prev}{rr}+{cl}{ccf_flujo_row}")
    set_num(ws, rr, 11, formula=f"=H{rr}")
    rr += 1

    set_text(ws, rr, 2, "Por recuperar inversor (SAFE)")
    for y in range(1, 6):
        c = y + 3
        cl = get_column_letter(c)
        set_num(ws, rr, c, formula=f"=MAX(0,-$C$9-{cl}{ccf_acum_flujo_row})")
    set_num(ws, rr, 12, formula=f"=H{rr}")
    rr += 1

    set_text(ws, rr, 2, "ECF (founders)")
    for y in range(1, 6):
        c = y + 3
        cl = get_column_letter(c)
        set_num(ws, rr, c, formula=f"={cl}{fcf_flujo_row}-{cl}{ccf_flujo_row}")
    rr += 2

    # Vector IRR en columna J (t=0 + FCF años 1–5)
    set_text(ws, rr, 10, "Flujo IRR")
    irr_start = rr + 1
    set_num(ws, irr_start, 10, value=-SAFE_LEAN)
    for y in range(1, 6):
        c = y + 3
        cl = get_column_letter(c)
        set_num(ws, irr_start + y, 10, formula=f"={cl}{fcf_flujo_row}")

    van_row = rr + 7
    set_text(ws, van_row, 2, "VP Flujos(5)")
    set_num(ws, van_row, 3, formula=f"=NPV($C$8,D{fcf_flujo_row}:H{fcf_flujo_row})")
    set_text(ws, van_row + 1, 2, "VAN(5)")
    set_num(ws, van_row + 1, 3, formula=f"=C{van_row}+C9")
    set_text(ws, van_row + 2, 2, "VP Flujos(3)")
    set_num(ws, van_row + 2, 3, formula=f"=NPV($C$8,D{fcf_flujo_row}:F{fcf_flujo_row})")
    set_text(ws, van_row + 3, 2, "VAN(3)")
    set_num(ws, van_row + 3, 3, formula=f"=C{van_row + 2}+C9")
    set_text(ws, van_row + 4, 2, "TIR(5)")
    set_num(ws, van_row + 4, 3, formula=f"=IRR(J{irr_start}:J{irr_start + 5})")
    set_text(ws, van_row + 5, 2, "TIR(3)")
    set_num(ws, van_row + 5, 3, formula=f"=IRR(J{irr_start}:J{irr_start + 3})")
    set_text(ws, van_row + 6, 2, "Payback inversor (años, ilustrativo)")
    set_num(
        ws, van_row + 6, 3,
        formula=(
            f'=IF(H{ccf_acum_flujo_row}>=ABS($C$9),5,'
            f'IF(G{ccf_acum_flujo_row}>=ABS($C$9),4,'
            f'IF(F{ccf_acum_flujo_row}>=ABS($C$9),3,'
            f'IF(E{ccf_acum_flujo_row}>=ABS($C$9),2,'
            f'IF(D{ccf_acum_flujo_row}>=ABS($C$9),1,"[LARGO PLAZO]")))))'
        ),
    )

    rr = van_row + 8
    set_text(
        ws,
        rr,
        2,
        f"NOTA: Reparto CCF/ECF ilustrativo post-cap SAFE {_equity_pct_es()}%. No cláusula de dividendos.",
    )

    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 14


def build_tasa_crecimiento(ws):
    ws.title = "Tasa Crecimiento"
    set_text(ws, 2, 2, "TASAS DE CRECIMIENTO — Zonix Pharma", True)
    for y in range(1, 6):
        set_text(ws, 4, y + 2, f"Año {y}")
    style_header_row(ws, 4, range(3, 8))
    r = 6

    set_text(ws, r, 2, "Revenue anual % vs año anterior")
    for y in range(2, 6):
        c = y + 2
        prev = get_column_letter(c - 1)
        cl = get_column_letter(c)
        set_num(ws, r, c, formula=f"=IF({prev}{r+2}=0,\"\",({cl}{r+2}-{prev}{r+2})/{prev}{r+2})", fmt=PCT_FMT)
    r += 1

    set_text(ws, r, 2, "Farmacias activas cierre")
    for y in range(1, 6):
        ref = YEAR_ROWS[y]["activas"]
        set_num(ws, r, y + 2, formula=f"='Año {y}'!N{ref}")
    r += 1

    set_text(ws, r, 2, "Revenue anual USD")
    for y in range(1, 6):
        ref = YEAR_ROWS[y]["ing"]
        set_num(ws, r, y + 2, formula=f"='Año {y}'!O{ref}")
    r += 1

    set_text(ws, r, 2, "ARPF placeholder USD")
    arpf_vals = [50, 52, 55, 58, 60]
    for j, v in enumerate(arpf_vals):
        set_num(ws, r, j + 3, value=v)
    r += 1

    set_text(ws, r, 2, "Burn prom. mensual USD")
    for y in range(1, 6):
        ref = YEAR_ROWS[y]["cost"]
        set_num(ws, r, y + 2, formula=f"='Año {y}'!O{ref}/12")
    r += 1

    set_text(ws, r, 2, "Etiqueta")
    for j, tag in enumerate(["ancla", "PROY §2", "PROY §3", "SUPUESTO", "SUPUESTO"]):
        set_text(ws, r, j + 3, tag)


def finish_workbook(wb: Workbook):
    for name in wb.sheetnames:
        ws = wb[name]
        if name.startswith("Año"):
            ws.freeze_panes = "C16" if name == "Año 1" else "C11"
            ws.sheet_view.zoomScale = 90
        if name == "Detallado de la inversión.":
            ws.freeze_panes = "C10"


def main():
    global YEAR_ROWS, DETALLADO_ROWS, ESTA_BURN_ROWS
    YEAR_ROWS = {}
    DETALLADO_ROWS = {}
    ESTA_BURN_ROWS = {}

    wb = Workbook()
    wb.remove(wb.active)

    build_detallado(wb.create_sheet("Detallado de la inversión."))
    DETALLADO_ROWS["burn_row"] = next(
        r for r in range(1, 200)
        if wb["Detallado de la inversión."].cell(r, 3).value == "Burn M1-M12"
    )
    DETALLADO_ROWS["reserva_row"] = DETALLADO_ROWS["burn_row"] + 1

    build_hoja3(wb.create_sheet("Hoja3"))
    build_hoja1(wb.create_sheet("Hoja1"))
    build_hoja2(wb.create_sheet("Hoja2"))
    build_esta_si_vale(wb.create_sheet("ESTA SI VALE"))

    for y in range(1, 6):
        build_year_sheet(wb.create_sheet(f"Año {y}"), y, wb)

    populate_esta_simulator(wb["ESTA SI VALE"], wb)

    build_flujo_total(wb)
    build_tasa_crecimiento(wb.create_sheet("Tasa Crecimiento"))
    finish_workbook(wb)

    wb.save(OUT)
    print(f"Written: {OUT}")
    print(
        f"{MODEL_VERSION} anclas Lean: SAFE={SAFE_LEAN:,} Fase0={FASE0_TOTAL:,} Day-D={CAJA_DAYD:,} "
        f"BurnY1={BURN_Y1:,} RevY1={REVENUE_Y1:,} FCF={FCF_Y1:,} CashM12≈{round(CAJA_M12_THEORETICAL):,}"
        .replace(",", ".")
    )
    for key, s in TIER_SUMMARY.items():
        print(
            f"  Tier {key}: capital={s['capital']:,} burn_avg={s['burn_avg']:,} "
            f"fase0={s['fase0']:,} safe_cap={s['safe_cap']:,}".replace(",", ".")
        )
    try:
        import shutil

        shutil.copy2(OUT, DESCARGAS)
        print(f"Copied: {DESCARGAS}")
    except OSError as e:
        print(f"Warning: could not copy to Descargas: {e}")
        print(f"  Manual: cp {OUT} ~/Descargas/")


if __name__ == "__main__":
    main()
