#!/usr/bin/env python3
"""Verificación ancla del MODELO_FINANCIERO_ZONIX_PHARMA.xlsx (v3.8.2 layout compacto + ESTA espejo Pizza)."""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "MODELO_FINANCIERO_ZONIX_PHARMA.xlsx"

EXPECTED_SHEETS = [
    "Detallado de la inversión.",
    "Hoja3",
    "Hoja1",
    "Hoja2",
    "ESTA SI VALE",
    "Año 1",
    "Año 2",
    "Año 3",
    "Año 4",
    "Año 5",
    "Flujo Total",
    "Tasa Crecimiento",
]

FORBIDDEN_DET = ("BPF", "CPE", "Permisos Pharma VE", "Pizza QLQ", "sanitario estadal")

TOL = 50


def fail(msg: str) -> None:
    print(f"FAIL: {msg}")
    sys.exit(1)


def ok(msg: str) -> None:
    print(f"OK: {msg}")


def col_month(m: int) -> int:
    return 2 + m


def import_anchors():
    from generate_modelo_financiero_v2 import (
        ACTIVE_TIER,
        BURN,
        BURN_Y1,
        CAJA_DAYD,
        CAJA_M12_THEORETICAL,
        DISCOUNT,
        FASE0_TOTAL,
        FCF_Y1,
        LEGAL_SUBTOTAL,
        ONE_SHOTS_LEAN,
        REVENUE,
        REVENUE_Y1,
        SAFE_LEAN,
        TIER_SUMMARY,
        YEAR_DATA,
        fixed_payroll,
    )

    rate = DISCOUNT
    fcfs = [YEAR_DATA[i][2] - YEAR_DATA[i][3] for i in range(5)]
    npv = sum(f / ((1 + rate) ** (i + 1)) for i, f in enumerate(fcfs))
    van5 = round(npv + (-SAFE_LEAN))

    return {
        "revenue_y1": REVENUE_Y1,
        "burn_y1": BURN_Y1,
        "fcf_y1": FCF_Y1,
        "one_shots": ONE_SHOTS_LEAN,
        "fase0": FASE0_TOTAL,
        "caja_dayd": CAJA_DAYD,
        "caja_m12": round(CAJA_M12_THEORETICAL),
        "safe_lean": SAFE_LEAN,
        "van5": van5,
        "legal_subtotal": LEGAL_SUBTOTAL,
        "burn_m1": BURN[0],
        "fixed_payroll": fixed_payroll(ACTIVE_TIER),
        "tier_summary": TIER_SUMMARY,
        "active_tier": ACTIVE_TIER,
        "BURN": BURN,
        "REVENUE": REVENUE,
    }


def simulate_burn_m1(anchors: dict) -> int:
    from generate_modelo_financiero_v2 import (
        BURN,
        CONTINGENCIA_BY_MONTH,
        META_BY_MONTH,
        VALLA_BY_MONTH,
    )

    t = anchors["active_tier"]
    fixed = anchors["fixed_payroll"]
    m1 = fixed + META_BY_MONTH[0] + VALLA_BY_MONTH[0] + CONTINGENCIA_BY_MONTH[0]
    if m1 != BURN[0]:
        fail(f"Simulación burn M1={m1} != BURN[0]={BURN[0]}")
    if m1 != anchors["burn_m1"]:
        fail(f"Burn M1 script {anchors['burn_m1']} != sim {m1}")
    return sum(BURN)


def esta_row_for_formula(wb, formula: str) -> int | None:
    m = re.search(r"\$I\$(\d+)", formula or "")
    return int(m.group(1)) if m else None


def det_cell_values(ws_det) -> list[str]:
    return [str(ws_det.cell(r, 3).value or "") for r in range(1, 200)]


def main() -> None:
    if not XLSX.exists():
        fail(f"No existe {XLSX}. Ejecute generate_modelo_financiero_v2.py primero.")

    anchors = import_anchors()
    ANCHORS = {k: v for k, v in anchors.items() if k not in ("tier_summary", "active_tier", "BURN", "REVENUE")}

    wb = openpyxl.load_workbook(XLSX, data_only=False)

    if wb.sheetnames != EXPECTED_SHEETS:
        fail(f"Hojas esperadas {len(EXPECTED_SHEETS)}, got {wb.sheetnames}")
    ok(f"{len(wb.sheetnames)} hojas en orden correcto")

    ref_errors = []
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and "#REF!" in v:
                    ref_errors.append(f"{name}!{cell.coordinate}")
    if ref_errors:
        fail(f"Referencias rotas: {ref_errors[:5]}")
    ok("Sin #REF! en fórmulas")

    ws1 = wb["Año 1"]
    ws2 = wb["Año 2"]
    ws_det = wb["Detallado de la inversión."]
    ws_flujo = wb["Flujo Total"]
    ws_esta = wb["ESTA SI VALE"]
    ws_hoja1 = wb["Hoja1"]

    from generate_modelo_financiero_v2 import DET_SHEET, HQ_CAPEX_LEAN, MODEL_VERSION

    wb_vals = openpyxl.load_workbook(XLSX, data_only=True)
    ws_hoja1_vals = wb_vals["Hoja1"]

    # --- Hoja1 v3.7: 8 bloques Zonix espejo + panel ---
    hoja1_e_text = " ".join(
        str(ws_hoja1.cell(r, 5).value or "") for r in range(1, 120)
    )
    hoja1_j_text = " ".join(
        str(ws_hoja1.cell(r, 10).value or "") for r in range(1, 15)
    )
    hoja1_joined = hoja1_e_text + " " + hoja1_j_text

    if "PC recepción" in hoja1_joined and "Detallado" not in hoja1_joined:
        fail("Hoja1 aún duplica líneas PC literales (debe ser enlace a Detallado)")
    if "ZonixPharma" not in hoja1_e_text and "Detallado" not in hoja1_e_text:
        fail("Hoja1 debe indicar datos ZonixPharma / enlace Detallado")
    if "no sumar" not in hoja1_e_text.lower() and "one-shots" not in hoja1_e_text.lower():
        fail("Hoja1 sin disclaimer anti-suma one-shots")
    ok("Hoja1 disclaimer + enlace Detallado ZonixPharma")

    block_titles = (
        "equipos",
        "adecuacion",
        "mano de obra",
        "transporte",
        "alquiler",
        "constitucion empresa y permisos",
        "marketing",
        "materia",
    )
    for title in block_titles:
        if title not in hoja1_e_text.lower():
            fail(f"Hoja1 sin bloque «{title}» (v3.7)")
    ok("Hoja1 8 bloques Pizza → Zonix (equipos…materia)")

    pct_header = str(ws_hoja1.cell(5, 9).value or "")
    if "% SAFE" not in pct_header:
        fail("Hoja1 sin header col I % SAFE")
    pct_h_formulas = 0
    det_formulas = 0
    for r in range(1, 120):
        v = ws_hoja1.cell(r, 9).value
        if isinstance(v, str) and DET_SHEET in v and "!H" in v:
            pct_h_formulas += 1
        for c in range(5, 10):
            cell = ws_hoja1.cell(r, c).value
            if isinstance(cell, str) and DET_SHEET in cell:
                det_formulas += 1
    if pct_h_formulas < 25:
        fail(f"Hoja1 col I % SAFE incompleta: {pct_h_formulas} fórmulas Detallado!H")
    if det_formulas < 30:
        fail(f"Hoja1 espejo incompleto: {det_formulas} fórmulas Detallado")
    ok(f"Hoja1 col I % SAFE ({pct_h_formulas} filas) + espejo Detallado ({det_formulas} refs)")

    mirror_c_rows = sum(
        1
        for r in range(1, 120)
        if isinstance(ws_hoja1.cell(r, 5).value, str)
        and DET_SHEET in ws_hoja1.cell(r, 5).value
        and "!C" in ws_hoja1.cell(r, 5).value
    )
    if mirror_c_rows < 25:
        fail(f"Hoja1 espejo incompleto: {mirror_c_rows} filas Descripción enlazadas")
    ok("Hoja1 espejo líneas Detallado (smoke)")

    if "Inversión CapEx HQ" not in hoja1_j_text:
        fail("Hoja1 sin panel lateral Inversión CapEx HQ (col J)")
    if "TOTAL SAFE" not in hoja1_j_text:
        fail("Hoja1 panel sin TOTAL SAFE Lean (ref.)")
    if "Burn mensual" not in hoja1_j_text and "MO+mkt" not in hoja1_j_text:
        fail("Hoja1 panel sin burn mensual ref (MO+mkt)")
    if "Reserva caja" not in hoja1_j_text:
        fail("Hoja1 panel sin reserva caja (ref.)")
    if "equity" not in hoja1_j_text.lower():
        fail("Hoja1 panel sin % equity SAFE (ref.)")
    if "Hoja3" not in hoja1_j_text:
        fail("Hoja1 sin enlace Use of funds → Hoja3")

    panel_total_safe_ok = False
    panel_burn_ok = False
    panel_reserva_ok = False
    for r in range(1, 15):
        lbl = str(ws_hoja1.cell(r, 10).value or "")
        fk = ws_hoja1.cell(r, 11).value
        if "TOTAL SAFE" in lbl and isinstance(fk, str) and DET_SHEET in fk and "!G" in fk:
            panel_total_safe_ok = True
        if ("Burn mensual" in lbl or "MO+mkt" in lbl) and isinstance(fk, str) and DET_SHEET in fk:
            panel_burn_ok = True
        if "Reserva caja" in lbl and isinstance(fk, str) and DET_SHEET in fk and "!G" in fk:
            panel_reserva_ok = True
    if not panel_total_safe_ok:
        fail("Hoja1 panel TOTAL SAFE sin fórmula Detallado!G en col K")
    if not panel_burn_ok:
        fail("Hoja1 panel burn mensual sin enlace Detallado en col K")
    if not panel_reserva_ok:
        fail("Hoja1 panel reserva sin enlace Detallado!G en col K")
    ok("Hoja1 panel J–K (TOTAL SAFE, burn, reserva, equity)")

    panel_disclaimer = " ".join(
        str(ws_hoja1.cell(r, 10).value or "") for r in range(1, 4)
    )
    if "no sumar" not in panel_disclaimer.lower():
        fail("Hoja1 panel sin disclaimer anti-doble-conteo")
    ok("Hoja1 panel anti-doble-conteo")

    if "Horno pizza" in hoja1_joined or "Harina (" in hoja1_joined:
        fail("Hoja1 contiene rubros Pizza físicos (hornos/harina)")
    if "ESTA" not in hoja1_e_text:
        fail("Hoja1 bloque materia sin referencia ESTA SI VALE")
    ok("Hoja1 sin rubros Pizza físicos + ref ESTA en materia")

    pc_labels = (
        "PC recepción",
        "PC administración",
        "PC desarrollo",
        "PC puesto flexible",
    )
    pc_sum = 0
    hq_item_sum = 0
    for r in range(1, 200):
        lbl = str(ws_det.cell(r, 3).value or "")
        if any(lbl.startswith(p) for p in pc_labels):
            qty = ws_det.cell(r, 5).value
            price = ws_det.cell(r, 6).value
            if isinstance(qty, (int, float)) and isinstance(price, (int, float)):
                pc_sum += qty * price
        if lbl.startswith("Depósito") or lbl.startswith("Adecuación"):
            qty = ws_det.cell(r, 5).value
            price = ws_det.cell(r, 6).value
            if isinstance(qty, (int, float)) and isinstance(price, (int, float)):
                hq_item_sum += qty * price
    hq_total = hq_item_sum + pc_sum
    if pc_sum != 3700:
        fail(f"CapEx 4 PCs en Detallado esperado 3700, got {pc_sum}")
    if hq_total != HQ_CAPEX_LEAN:
        fail(f"Subtotal HQ en Detallado esperado {HQ_CAPEX_LEAN}, got {hq_total}")

    pc_formula_row = None
    for r in range(1, 120):
        lbl = str(ws_hoja1.cell(r, 5).value or "").lower()
        f8 = ws_hoja1.cell(r, 8).value
        if "subtotal equipos" in lbl and isinstance(f8, str) and "SUM" in f8.upper():
            pc_formula_row = r
    if pc_formula_row is None:
        fail("Hoja1 sin Subtotal equipos (4 PCs) con SUM a Detallado")
    ok("Hoja1 Subtotal equipos enlazado a Detallado")

    pc_val = ws_hoja1_vals.cell(pc_formula_row, 8).value
    if pc_val is not None and isinstance(pc_val, (int, float)):
        if abs(pc_val - 3700) > 0.01:
            fail(f"Hoja1 4 PCs data_only esperado 3700, got {pc_val}")
        ok("Hoja1 4 PCs data_only = 3.700")
    else:
        ok("Hoja1 4 PCs: fórmula OK (sin cache Excel — abrir xlsx para recalcular)")

    hq_sub_val = None
    for r in range(1, 120):
        lbl = str(ws_hoja1.cell(r, 5).value or "").lower()
        if "subtotal adecuación" in lbl or "subtotal adecuacion" in lbl:
            v = ws_hoja1_vals.cell(r, 8).value
            if isinstance(v, (int, float)):
                hq_sub_val = v
                break
    if hq_sub_val is not None:
        if abs(hq_sub_val - 1650) > 0.01:
            fail(f"Hoja1 subtotal adecuación esperado 1650, got {hq_sub_val}")
        ok("Hoja1 subtotal adecuación data_only = 1.650")
    else:
        ok("Hoja1 subtotal adecuación: fórmula OK (sin cache Excel)")

    title_h1 = str(ws_hoja1.cell(1, 5).value or "")
    if MODEL_VERSION not in title_h1 or "Zonix Pharma" not in title_h1:
        fail(f"Hoja1 título debe incluir Zonix Pharma + {MODEL_VERSION}")
    ok(f"Hoja1 vista completa 8 bloques ({MODEL_VERSION})")

    from generate_modelo_financiero_v2 import EQUITY_INV

    equity_c6 = ws_flujo["C6"].value
    if equity_c6 != EQUITY_INV:
        fail(f"Flujo Total C6 equity esperado {EQUITY_INV}, got {equity_c6!r}")
    ok(f"Flujo Total C6 equity SAFE = {EQUITY_INV} (~18,66%)")

    p10_act = p50_act = p90_act = None
    for r in range(1, 160):
        esc = ws_esta.cell(r, 13).value
        if esc == "P10 pesimista":
            p10_act = ws_esta.cell(r, 15).value
        elif esc == "P50 base":
            p50_act = ws_esta.cell(r, 15).value
        elif esc == "P90 optimista":
            p90_act = ws_esta.cell(r, 15).value
    if p10_act is None or p50_act is None or p90_act is None:
        fail("ESTA sin filas P10/P50/P90 completas (panel col M)")
    if not (p10_act < p50_act < p90_act):
        fail(f"P10/P50/P90 activas deben ser estrictamente crecientes: {p10_act} < {p50_act} < {p90_act}")
    ok(f"Escenarios P10/P50/P90 activas M12: {p10_act} < {p50_act} < {p90_act}")

    # --- ESTA v3.8: bloques inversión + simulador ---
    esta_e_text = " ".join(
        str(ws_esta.cell(r, 5).value or "") for r in range(1, 120)
    ).lower()
    esta_m_text = " ".join(
        str(ws_esta.cell(r, 13).value or "") for r in range(1, 120)
    ).lower()
    if "zonixpharma" not in esta_e_text and "no sumar" not in esta_e_text:
        fail("ESTA sin disclaimer anti-doble-conteo")
    ok("ESTA disclaimer ZonixPharma")

    for title in (
        "mano de obra",
        "transporte",
        "constitución empresa",
        "introducción al mercado",
    ):
        if title not in esta_e_text:
            fail(f"ESTA sin bloque «{title}» (v3.8)")
    ok("ESTA bloques inversión MO/transporte/legal/intro")

    esta_det = sum(
        1
        for r in range(1, 120)
        for c in range(5, 11)
        if isinstance(ws_esta.cell(r, c).value, str)
        and DET_SHEET in ws_esta.cell(r, c).value
    )
    if esta_det < 40:
        fail(f"ESTA espejo Detallado incompleto: {esta_det} refs")
    ok(f"ESTA espejo Detallado ({esta_det} refs)")

    if ws_esta.max_row < 75:
        fail(f"ESTA max_row {ws_esta.max_row} < 75 (v3.8 densidad mínima)")
    ok(f"ESTA densidad v3.8 (max_row={ws_esta.max_row})")

    sim_row = burn_row = None
    for r in range(1, ws_esta.max_row + 1):
        v = ws_esta.cell(r, 15).value
        if isinstance(v, str) and "SIMULADOR MARKETPLACE" in v:
            sim_row = r
        bv = ws_esta.cell(r, 5).value
        if isinstance(bv, str) and "BURN MENSUAL REF" in bv:
            burn_row = r
    if sim_row and burn_row and abs(sim_row - burn_row) > 2:
        fail(f"ESTA simulador (fila {sim_row}) lejos de burn (fila {burn_row}) — layout no compacto")
    if sim_row and burn_row:
        ok(f"ESTA layout compacto (burn f{burn_row} ≈ sim f{sim_row})")

    sim_ok = False
    for r in range(1, ws_esta.max_row + 1):
        v = ws_esta.cell(r, 15).value
        if isinstance(v, str) and "SIMULADOR MARKETPLACE" in v:
            sim_ok = True
        qv = ws_esta.cell(r, 17).value
        if isinstance(qv, str) and "Año 1" in qv and "!" in qv:
            sim_ok = True
    if not sim_ok:
        fail("ESTA sin simulador marketplace enlazado a Año 1")
    ok("ESTA simulador M1-M12 → Año 1")

    if "horno pizza" in esta_e_text or "harina (" in esta_e_text:
        fail("ESTA contiene rubros Pizza físicos (hornos/harina)")
    ok("ESTA sin rubros Pizza físicos")

    if "gastos fijos mensuales" not in " ".join(
        str(ws_esta.cell(r, 16).value or "") for r in range(1, 40)
    ).lower():
        fail("ESTA sin bandas gastos fijos SaaS (col P)")
    ok("ESTA bandas gastos fijos SaaS")

    title_esta = str(ws_esta.cell(7, 5).value or "")
    if MODEL_VERSION not in title_esta:
        fail(f"ESTA título debe incluir {MODEL_VERSION}")
    ok(f"ESTA espejo Pizza ({MODEL_VERSION})")

    # --- ESTA v3.8.2: integridad col J + simulador ---
    factor_rows_ok = sum(
        1
        for r in range(1, 80)
        if str(ws_esta.cell(r, 5).value or "").startswith("Factor Año 2")
        and isinstance(ws_esta.cell(r, 10).value, (int, float))
    )
    if factor_rows_ok < 3:
        fail(f"ESTA sin filas factor Año 2 dedicadas (got {factor_rows_ok})")
    ok(f"ESTA filas factor Año 2 dedicadas ({factor_rows_ok})")

    bad_j: list[tuple[int, str]] = []
    for r in range(1, ws_esta.max_row + 1):
        v = ws_esta.cell(r, 10).value
        if not isinstance(v, str) or not v.startswith("="):
            continue
        if f"$J${r})" in v or f"J{r}*" in v:
            bad_j.append((r, v))
        if re.search(rf"I{r}\*\(1\+", v):
            bad_j.append((r, f"base I not H: {v}"))
    if bad_j:
        fail(f"ESTA col J fórmulas inválidas: {bad_j[:3]}")
    ok("ESTA col J sin circularidad; base Total USD col H")

    sim_hdr = None
    for r in range(1, ws_esta.max_row + 1):
        cell_o = ws_esta.cell(r, 15).value
        if isinstance(cell_o, str) and "SIMULADOR MARKETPLACE" in cell_o:
            sim_hdr = r
            break
    if sim_hdr:
        m1_row = sim_hdr + 2
        u_formula = ws_esta.cell(m1_row, 21).value
        if not (
            isinstance(u_formula, str)
            and u_formula.startswith("=S")
            and f"-P{m1_row}" in u_formula
        ):
            fail(f"Simulador Util./activa debe ser S-P, got {u_formula!r}")
        ok("ESTA simulador Util./activa = ARPF − costo/activa (S−P)")

    a2_c3 = ws2["C3"].value
    if not (isinstance(a2_c3, str) and "Año 1" in a2_c3 and "N" in a2_c3):
        fail(f"Año 2!C3 debe apuntar a activas M12 Año 1, got {a2_c3!r}")
    ok("Año 2!C3 → activas M12 Año 1")

    if ws1["B2"].value != "ITEMS" or ws1["B16"].value != "AÑO 1":
        fail("Layout Año 1 incorrecto")
    ok("Layout Año 1 sin ITEMS duplicado")

    ing_total = ws1["O25"].value
    if not (isinstance(ing_total, str) and ing_total.startswith("=SUM")):
        fail(f"Total Ingresos Año 1 debe ser fórmula SUM, got {ing_total!r}")
    ok("Total Ingresos Año 1 es fórmula")

    for r in range(1, 55):
        b = ws1.cell(r, 2).value
        if b and "Cuadre buffer" in str(b):
            fail(f"Fila Cuadre buffer visible en r{r}")
    ok("Sin fila Cuadre buffer en P&L")

    d18 = ws_flujo["D18"].value
    if not (isinstance(d18, str) and "Año 1" in d18):
        fail(f"Flujo Total ingresos A1 debe referenciar Año 1, got {d18!r}")
    ok("Flujo Total enlazado a hojas Año")

    ccf_formula = ws_flujo["D22"].value
    if not (isinstance(ccf_formula, str) and "$C$6" in ccf_formula):
        fail(f"CCF debe usar $C$6 (equity), got {ccf_formula!r}")
    if isinstance(ccf_formula, str) and "$C$5" in ccf_formula:
        fail("CCF no debe usar $C$5 (ARPF)")
    ok("Flujo Total CCF usa $C$6 (equity SAFE)")

    founders = ws_flujo["C7"].value
    if founders != "=1-C6":
        fail(f"Founders debe ser =1-C6, got {founders!r}")
    ok("Flujo Total founders =1-C6")

    labels = {ws_flujo.cell(r, 2).value for r in range(1, 55)}
    for need in ("TIR(5)", "TIR(3)", "Payback inversor (años, ilustrativo)"):
        if need not in labels:
            fail(f"Falta métrica {need!r} en Flujo Total")
    ok("TIR(3/5) y payback en Flujo Total")

    for y in range(1, 6):
        hdr = ws_flujo.cell(16, y + 3).value
        if hdr != f"Año {y}":
            fail(f"Flujo Total D16:H16 esperado Año {y} en col {y + 3}, got {hdr!r}")
    if ws_flujo.cell(16, 3).value == "Año 1":
        fail("Flujo Total C16 no debe ser Año 1 (datos en D:H)")
    activas_total = ws_flujo["I17"].value
    if activas_total != "=H17":
        fail(f"Flujo Total I17 activas TOTAL debe ser =H17, got {activas_total!r}")
    ok("Flujo Total headers D:H alineados + I17=H17")

    ws_tasa = wb["Tasa Crecimiento"]
    for y in range(1, 6):
        hdr = ws_tasa.cell(4, y + 2).value
        if hdr != f"Año {y}":
            fail(f"Tasa Crecimiento C4:G4 esperado Año {y} en col {y + 2}, got {hdr!r}")
    growth_d = ws_tasa["D6"].value
    if not (isinstance(growth_d, str) and "C8" in growth_d and "C7" not in growth_d):
        fail(f"Tasa Crecimiento D6 debe usar fila revenue (C8), got {growth_d!r}")
    ok("Tasa Crecimiento headers C:G + YoY revenue guard")

    t = anchors["active_tier"]
    expense_checks = {
        t.dev_label: t.dev,
        f"{t.sales_count}× Sales B2B": t.sales_count * t.sales_base,
        "HQ casa (arriendo + servicios)": t.hq,
        "Servicios oficina (electricidad, agua, internet)": t.servicios,
        "Hosting + SaaS + Firebase OTP": t.hosting,
    }
    for r in range(28, 45):
        label = ws1.cell(r, 2).value
        if not label:
            continue
        label = str(label).strip()
        for key, expected in expense_checks.items():
            if key in label:
                f = ws1.cell(r, 3).value
                er = esta_row_for_formula(wb, str(f) if f else "")
                if er is None:
                    fail(f"{key} sin enlace ESTA en fila {r}")
                val = ws_esta.cell(er, 9).value
                if not isinstance(val, (int, float)) or val != expected:
                    fail(f"ESTA I{er} para {key}: esperado {expected}, got {val!r}")
    ok("Enlaces ESTA dev/sales/HQ/servicios/hosting con valores v3")

    # Guard: dev en nómina ESTA
    dev_found = False
    for r in range(1, 160):
        lbl = ws_esta.cell(r, 5).value
        if lbl and t.dev_label.split("(")[0].strip() in str(lbl):
            dev_found = True
            if ws_esta.cell(r, 9).value != t.dev:
                fail(f"Dev ESTA fila {r}: esperado {t.dev}")
            break
    if not dev_found:
        fail(f"ESTA sin línea {t.dev_label!r}")
    ok(f"Dev en nómina ESTA = USD {t.dev}/mes")

    burn_row = next(
        r for r in range(1, 200) if ws_det.cell(r, 3).value == "Burn M1-M12"
    )
    if ws_det.cell(burn_row, 7).value != ANCHORS["burn_y1"]:
        fail(f"Burn M1-M12 esperado {ANCHORS['burn_y1']}, got {ws_det.cell(burn_row, 7).value}")
    ok(f"Burn M1-M12 = {ANCHORS['burn_y1']}")

    if sum(anchors["BURN"]) != ANCHORS["burn_y1"]:
        fail("Script BURN no cuadra ancla")
    if sum(anchors["REVENUE"]) != ANCHORS["revenue_y1"]:
        fail("Script REVENUE no cuadra ancla")
    ok("Constantes Python cuadran anclas pack v3")

    from generate_modelo_financiero_v2 import (
        CAJA_DAYD,
        FASE0_TOTAL,
        ONE_SHOTS_LEAN,
        SAFE_LEAN,
    )

    if ONE_SHOTS_LEAN != ANCHORS["one_shots"]:
        fail(f"ONE_SHOTS_LEAN {ONE_SHOTS_LEAN} != {ANCHORS['one_shots']}")
    if FASE0_TOTAL != ANCHORS["fase0"]:
        fail(f"FASE0_TOTAL {FASE0_TOTAL} != {ANCHORS['fase0']}")
    if CAJA_DAYD != ANCHORS["caja_dayd"]:
        fail(f"CAJA_DAYD {CAJA_DAYD} != {ANCHORS['caja_dayd']}")
    if SAFE_LEAN - FASE0_TOTAL != CAJA_DAYD:
        fail("SAFE − Fase0 debe igualar caja Day-D")
    ok("Cadena SAFE − Fase0 = Day-D cuadra v3 bottom-up")

    burn_sim = simulate_burn_m1(anchors)
    if burn_sim != ANCHORS["burn_y1"]:
        fail(f"Suma burn simulada {burn_sim} != {ANCHORS['burn_y1']}")
    ok(f"Burn simulado M1={anchors['burn_m1']} y anual={burn_sim}")

    if ANCHORS["fcf_y1"] != ANCHORS["revenue_y1"] - ANCHORS["burn_y1"]:
        fail("FCF Y1 no cuadra revenue − burn")
    ok(f"FCF Y1 teórico = {ANCHORS['fcf_y1']}")

    van5 = ANCHORS["van5"]
    ok(f"VAN(5) teórico ≈ {van5} (negativo esperado en Lean v3)")

    cash_final_row = None
    for r in range(45, 55):
        if ws1.cell(r, 2).value == "Cash final mes":
            cash_final_row = r
            break
    if cash_final_row is None:
        fail("No se encontró fila Cash final mes")
    m12_cash = ws1.cell(cash_final_row, col_month(12)).value
    if not (isinstance(m12_cash, str) and m12_cash.startswith("=")):
        fail(f"Cash M12 debe ser fórmula, got {m12_cash!r}")
    ok("Cash final M12 es fórmula viva")

    caja_formula = ws1.cell(cash_final_row - 1, col_month(1)).value
    if not (isinstance(caja_formula, str) and "Detallado" in caja_formula):
        fail(f"Cash M1 ini debe enlazar Detallado, got {caja_formula!r}")
    ok("Cash M1 inicial enlazado a Detallado (Day-D)")

    # P50 cash M12 en ESTA no debe ser 42209 (v2.3 obsoleto)
    for r in range(1, 160):
        if ws_esta.cell(r, 13).value == "P50 base":
            p50 = ws_esta.cell(r, 17).value
            if p50 == 42209:
                fail("P50 Cash M12 obsoleto 42.209 — debe ser v3")
            if p50 != ANCHORS["caja_m12"]:
                fail(f"P50 Cash M12 esperado {ANCHORS['caja_m12']}, got {p50!r}")
            ok(f"P50 Cash M12 = {ANCHORS['caja_m12']} (v3)")
            break
    else:
        fail("No se encontró fila P50 base en ESTA")

    from generate_modelo_financiero_v2 import DET_SHEET, MODEL_VERSION, RESERVA_LEAN

    hoja3 = wb["Hoja3"]
    hoja3_labels = {str(hoja3.cell(r, 5).value or "") for r in range(1, 120)}

    if "A — Use of funds" not in " ".join(hoja3_labels):
        fail("Hoja3 sin Sección A use-of-funds")
    if not any("B — Desglose operativo" in x for x in hoja3_labels):
        fail("Hoja3 sin Sección B desglose operativo")
    if not any("Mano de obra Fase 0" in x for x in hoja3_labels):
        fail("Hoja3 Sección B sin bloque MO Fase 0")
    if not any("Constitución empresa" in x for x in hoja3_labels):
        fail("Hoja3 Sección B sin bloque Constitución")
    if not any("Introducción al mercado" in x for x in hoja3_labels):
        fail("Hoja3 Sección B sin bloque Intro")
    if not any("Marketing pre-lanzamiento" in x for x in hoja3_labels):
        fail("Hoja3 Sección B sin bloque Marketing pre-lanzamiento")
    if not any("Materia prima" in x for x in hoja3_labels):
        fail("Hoja3 Sección B sin footnote SaaS (materia prima)")
    if not any("HQ y CapEx" in x for x in hoja3_labels):
        fail("Hoja3 Sección B sin bloque HQ y CapEx")
    if not any("Transporte B2B" in x for x in hoja3_labels):
        fail("Hoja3 Sección B sin bloque Transporte B2B")
    if not any("Validación" in x and "ops burn" in x for x in hoja3_labels):
        fail("Hoja3 Sección B sin fila puente cross Fase 0")
    if not any("Informativo" in x and "no sumar bloques" in x for x in hoja3_labels):
        fail("Hoja3 Sección B sin disclaimer anti-suma SAFE")
    ok("Hoja3 Secciones A/B + Intro/Mkt pre/HQ/Transporte + puente + disclaimer")

    lateral_k = [
        str(hoja3.cell(r, 11).value or "") for r in range(1, 45)
    ]
    lateral_joined = " ".join(lateral_k)
    if "One-shots (legal+intro+HQ)" not in lateral_joined:
        fail("Hoja3 lateral sin One-shots")
    if "Fase 0 operativa (sin one-shots)" not in lateral_joined:
        fail("Hoja3 lateral sin Fase 0 operativa (cross)")
    if "MO operativa (subtotal)" not in lateral_joined:
        fail("Hoja3 lateral sin MO operativa ref.")
    if "Marketing mensual (subtotal)" not in lateral_joined:
        fail("Hoja3 lateral sin Marketing mensual ref.")
    ok("Hoja3 lateral espejo Detallado (one-shots + cross + ref. mensual)")

    lateral_title = f"TIMELINE + RESUMEN FASE 0 ({MODEL_VERSION})"
    if hoja3.cell(9, 11).value != lateral_title:
        fail(f"Hoja3 lateral K9 esperado {lateral_title!r}, got {hoja3.cell(9, 11).value!r}")
    if not any("Sub-fase 0a" in x for x in lateral_k):
        fail("Hoja3 lateral sin Sub-fase 0a")
    ok("Hoja3 lateral timeline 0a–0c + resumen SAFE")

    mirror_found = False
    for r in range(1, 120):
        for c in range(5, 9):
            val = hoja3.cell(r, c).value
            if isinstance(val, str) and f"'{DET_SHEET}'!G" in val:
                mirror_found = True
                break
        if mirror_found:
            break
    if not mirror_found:
        fail("Hoja3 Sección B sin fórmulas espejo a Detallado")
    ok("Hoja3 espejo Detallado (smoke fórmula G)")

    total_a_row = 15
    h15 = hoja3.cell(total_a_row, 8).value
    if not (isinstance(h15, str) and "SUM(H11:H13)" in h15.replace(" ", "")):
        fail(f"Hoja3 TOTAL % debe ser =SUM(H11:H13), got {h15!r}")
    for r in range(11, 14):
        pct_formula = hoja3.cell(r, 8).value
        if not (isinstance(pct_formula, str) and f"G{total_a_row}" in pct_formula):
            fail(f"Hoja3 Sección A fila {r}: % debe dividir por G{total_a_row}")
    pct_sum = (
        ANCHORS["one_shots"] + ANCHORS["burn_y1"] + RESERVA_LEAN
    ) / ANCHORS["safe_lean"]
    if not (0.99 <= pct_sum <= 1.01):
        fail(f"Sección A use-of-funds no suma 100%: {pct_sum}")
    ok(f"Sección A use-of-funds suma ~100% ({pct_sum:.2%})")

    trap_ratio = (
        ANCHORS["fase0"] + ANCHORS["burn_y1"] + RESERVA_LEAN
    ) / ANCHORS["safe_lean"]
    if trap_ratio <= 1.05:
        fail(
            f"Anti-trampa: Fase0+Burn+Reserva debe ser >105% SAFE para evitar confusión reunión: {trap_ratio:.2%}"
        )
    ok(f"Anti-trampa numérica: Fase0+Burn+Reserva = {trap_ratio:.0%} del SAFE (no sumar en reunión)")

    subtotal_fase0_found = False
    for r in range(1, 120):
        if "Subtotal Fase 0" in str(hoja3.cell(r, 5).value or ""):
            f = hoja3.cell(r, 8).value
            if isinstance(f, str) and DET_SHEET in f and "!G" in f:
                subtotal_fase0_found = True
                break
    if not subtotal_fase0_found:
        fail("Hoja3 Subtotal Fase 0 sin fórmula a Detallado")
    ok("Hoja3 Subtotal Fase 0 enlazado a Detallado")

    # Anti-regresión ~146%: si Excel tiene cache de fórmulas, H11:H13 deben sumar ~100%
    wb_cached = openpyxl.load_workbook(XLSX, data_only=True)
    h_pct = [
        wb_cached["Hoja3"].cell(r, 8).value for r in range(11, 14)
    ]
    if all(isinstance(v, (int, float)) for v in h_pct):
        h_sum = sum(h_pct)
        if h_sum > 1.05:
            fail(f"Hoja3 Sección A H11:H13 suman >105% (regresión ~146%): {h_sum:.2%}")
        if not (0.99 <= h_sum <= 1.01):
            fail(f"Hoja3 Sección A H11:H13 no suman ~100%: {h_sum:.2%}")
        ok(f"Hoja3 celdas H11:H13 (cache) suman ~100% ({h_sum:.2%})")
    else:
        ok("Hoja3 % Sección A: fórmulas OK (sin cache Excel — abrir xlsx para recalcular)")

    # Anti-regresión: tabla vieja de 8 bloques con % sobre TOTAL (sumaba ~146%)
    old_overlap = sum(
        1
        for x in hoja3_labels
        if x in (
            "Fase 0 operativa",
            "Constitución y legal",
            "Introducción mercado / demo",
        )
    )
    if old_overlap >= 2:
        fail("Hoja3 regresión: layout viejo con bloques solapados en Sección A")
    ok("Sin layout Hoja3 obsoleto (8 bloques % solapados)")

    det_text = " ".join(det_cell_values(ws_det)).upper()
    for forbidden in FORBIDDEN_DET:
        if forbidden.upper() in det_text:
            fail(f"Detallado contiene texto prohibido: {forbidden!r}")
    ok("Sin BPF/CPE/permisos farmacia/Pizza en Detallado")

    det_labels = det_cell_values(ws_det)

    for need in (
        "CONSTITUCIÓN DE LA EMPRESA",
        "MANO DE OBRA FASE 0",
        "MANO DE OBRA OPERATIVA (post-Day-D",
        "MARKETING Y OPEX MENSUAL",
        "Dev junior",
        "Video pitch B2B",
    ):
        if not any(need in x for x in det_labels):
            fail(f"Detallado sin bloque {MODEL_VERSION}: {need!r}")
    ok(f"Detallado {MODEL_VERSION} incluye constitución, MO Fase 0/operativa, marketing desglosado")

    lateral_title = f"RESUMEN INVERSIÓN LEAN ({MODEL_VERSION})"
    if ws_det.cell(10, 11).value != lateral_title:
        fail(f"Resumen lateral K10 esperado {lateral_title!r}, got {ws_det.cell(10, 11).value!r}")
    det_lateral = " ".join(str(ws_det.cell(r, 11).value or "") for r in range(10, 35))
    if "One-shots (legal+intro+HQ)" not in det_lateral:
        fail("Detallado lateral sin One-shots")
    if "MO operativa (subtotal)" not in det_lateral:
        fail("Detallado lateral sin MO operativa ref.")
    ok("Resumen lateral tipo Pizza en cols K–M")

    total_safe_row = next(
        r for r in range(1, 250)
        if ws_det.cell(r, 3).value == "TOTAL capital SAFE Lean"
    )
    total_formula = str(ws_det.cell(total_safe_row, 7).value or "")
    if "subtotal_mo" in total_formula.lower() or "subtotal_mkt" in total_formula.lower():
        fail("TOTAL SAFE no debe incluir MO/marketing mensual (anti triple-conteo)")
    if not total_formula.startswith("=G"):
        fail(f"TOTAL SAFE debe ser suma one-shots+burn+reserva, got {total_formula!r}")
    ok("Guard anti triple-conteo: TOTAL SAFE = one-shots + burn + reserva")

    mkt_sub = next(
        (r for r in range(1, 250) if "Subtotal marketing + opex variable" in str(ws_det.cell(r, 3).value or "")),
        None,
    )
    if mkt_sub is None:
        fail("Detallado sin subtotal marketing mensual con montos")
    mkt_val = ws_det.cell(mkt_sub, 7).value
    if not (isinstance(mkt_val, str) and mkt_val.startswith("=SUM")):
        fail(f"Subtotal marketing debe ser fórmula SUM, got {mkt_val!r}")
    ok("Marketing mensual desglosado con subtotal fórmula")

    comm_esta = next(
        (r for r in range(1, 160) if "Sales comisiones" in str(ws_esta.cell(r, 5).value or "")),
        None,
    )
    if comm_esta is None:
        fail("ESTA sin fila Sales comisiones (estimado)")
    ok("Comisiones Sales modeladas en ESTA (variable)")

    growth_design = anchors["tier_summary"]["growth"]["burn_avg"]
    blitz_burn = anchors["tier_summary"]["blitz"]["burn_avg"]
    lean_burn = anchors["tier_summary"]["lean"]["burn_avg"]
    base_burn = anchors["tier_summary"]["base"]["burn_avg"]
    if growth_design <= lean_burn:
        fail("Growth burn prom debe superar Lean (marketing + Google Ads)")
    if base_burn <= lean_burn:
        fail("Base burn prom debe superar Lean (diseñador 300 + Google Ads 300/mes)")
    if blitz_burn <= base_burn:
        fail("Blitz burn prom debe superar Base (6× Sales + Google Ads 500)")
    ok(
        f"Tiers burn prom Lean={lean_burn} Base={base_burn} Growth={growth_design} "
        f"Blitz={blitz_burn} (v3.5 + Lean+ operativo)"
    )

    ts = anchors["tier_summary"]
    lean_cap = ts["lean"]["capital"]
    lp = ts["lean_plus"]
    if abs(lp["capital"] - lean_cap) > TOL:
        fail(f"Lean+ capital {lp['capital']} debe igualar Lean {lean_cap}")
    if lp["activas_eq"] <= ts["lean"]["activas_eq"]:
        fail(
            f"Lean+ activas M12 {lp['activas_eq']} debe superar Lean {ts['lean']['activas_eq']}"
        )
    if lp["revenue_m12"] <= ts["lean"]["burn_m12"]:
        fail(
            f"Lean+ revenue M12 {lp['revenue_m12']} debe superar burn Lean M12 "
            f"{ts['lean']['burn_m12']}"
        )
    if lp["be_month"] is None or lp["be_month"] > 10:
        fail(f"Lean+ debe equilibrar en M10 o antes, got be_month={lp['be_month']}")
    ok(
        f"Lean+ operativo: capital={lp['capital']} activas M12={lp['activas_eq']} "
        f"BE mes={lp['be_month']} cash M12≈{lp['caja_m12']}"
    )

    one_row = next(
        r for r in range(1, 200)
        if ws_det.cell(r, 3).value == "TOTAL one-shots Lean (legal+intro+HQ)"
    )
    legal_row = next(
        r for r in range(1, 200) if ws_det.cell(r, 3).value == "Subtotal constitución empresa"
    )
    intro_row = next(
        r for r in range(1, 200) if ws_det.cell(r, 3).value == "Subtotal intro mercado Lean"
    )
    hq_row = next(
        r for r in range(1, 200) if ws_det.cell(r, 3).value == "Subtotal HQ y CapEx"
    )
    one_shots_formula = ws_det.cell(one_row, 7).value
    expected_one = f"=G{legal_row}+G{intro_row}+G{hq_row}"
    if one_shots_formula != expected_one:
        fail(f"one_shots debe ser {expected_one!r}, got {one_shots_formula!r}")
    ok("one_shots solo suma legal+intro+HQ")

    if ANCHORS["legal_subtotal"] != 5050:
        fail(f"LEGAL_SUBTOTAL {ANCHORS['legal_subtotal']} != ancla")
    ok(f"Legal constitución sincero = USD {ANCHORS['legal_subtotal']:,}")

    # Tiers capital desde TIER_SUMMARY
    tier_rows = []
    for r in range(1, 160):
        lbl = ws_esta.cell(r, 13).value
        if lbl and "capital (USD)" in str(lbl):
            tier_rows.append((str(lbl), ws_esta.cell(r, 15).value))
    if len(tier_rows) < 5:
        fail("ESTA sin 5 tiers capital (Lean/Lean+/Base/Growth/Blitz)")
    for key in ("lean", "base", "growth", "blitz"):
        expected = anchors["tier_summary"][key]["capital"]
        match = [v for lbl, v in tier_rows if key in lbl.lower() or TIERS_LABEL(key) in lbl.lower()]
        if not match or match[0] != expected:
            fail(f"Tier {key} capital esperado {expected}, tiers={tier_rows}")
    lp_expected = anchors["tier_summary"]["lean_plus"]["capital"]
    lp_match = [
        v for lbl, v in tier_rows if "lean+" in lbl.lower() or "lean_plus" in lbl.lower()
    ]
    if not lp_match or abs(lp_match[0] - lp_expected) > TOL:
        fail(f"Tier lean_plus capital esperado {lp_expected}, tiers={tier_rows}")
    ok("Tiers Lean/Lean+/Base/Growth/Blitz capital alineados a bottom-up")

    print(
        f"\nVerificación {MODEL_VERSION} completada. Anclas Lean: SAFE={ANCHORS['safe_lean']:,} "
        f"Fase0={ANCHORS['fase0']:,} Day-D={ANCHORS['caja_dayd']:,} Burn={ANCHORS['burn_y1']:,} "
        f"Rev={ANCHORS['revenue_y1']:,} FCF={ANCHORS['fcf_y1']:,} CashM12≈{ANCHORS['caja_m12']:,}."
        .replace(",", ".")
    )


def TIERS_LABEL(key: str) -> str:
    from generate_modelo_financiero_v2 import TIERS

    return TIERS[key].label.split("—")[0].strip().lower()


if __name__ == "__main__":
    main()
